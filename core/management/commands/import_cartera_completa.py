from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from core.models import Cliente
import openpyxl
import os

class Command(BaseCommand):
    help = "Importar cartera completa de clientes desde Excel y asignar a managers"

    def handle(self, *args, **options):
        # Ruta del archivo Excel
        excel_path = '/home/crm_deploy/cartera_clientes/Cartera completa de clientes.xlsx'
        
        if not os.path.exists(excel_path):
            self.stdout.write(
                self.style.ERROR(f"✗ Archivo no encontrado: {excel_path}")
            )
            return

        # Mapeo de vendedores a usuarios managers
        vendedor_a_usuario = {
            'Rogers': 'rogers.orostica@gmail.com',
            'Magdalena': 'magdalena@claro.cl',
            'Carmen Gloria': 'carmen.gloria@claro.cl',
            'Cristian': 'cristian@claro.cl',
            'Lissette': 'lissette@claro.cl',
        }

        # Cargar Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Mapeo de ID usuario para asignación
        usuario_ids = {}
        for vendedor_nombre, email in vendedor_a_usuario.items():
            try:
                user = User.objects.get(email=email)
                usuario_ids[vendedor_nombre] = user.id
            except User.DoesNotExist:
                self.stdout.write(
                    self.style.WARNING(f"⚠ Usuario no encontrado: {email}")
                )

        # Contadores
        importados = 0
        duplicados = 0
        no_asignados = 0
        errores = 0

        # Importar clientes
        for row in range(2, ws.max_row + 1):
            try:
                rut = ws.cell(row, 1).value
                dv = ws.cell(row, 2).value
                razon_social = ws.cell(row, 3).value
                segmento = ws.cell(row, 4).value
                subrango = ws.cell(row, 5).value
                sub_movil = ws.cell(row, 6).value  # Vendedor
                subgerencia = ws.cell(row, 7).value
                renta_uf = ws.cell(row, 8).value
                q_total_lineas = ws.cell(row, 9).value
                bam = ws.cell(row, 10).value
                m2m = ws.cell(row, 11).value
                voz = ws.cell(row, 12).value
                fijo = ws.cell(row, 13).value
                movil = ws.cell(row, 14).value

                # Validar RUT
                if not rut:
                    continue

                # Convertir RUT a string y limpiar
                rut_str = str(int(rut)) if isinstance(rut, (int, float)) else str(rut)

                # Verificar si ya existe
                if Cliente.objects.filter(rut=rut_str).exists():
                    duplicados += 1
                    continue

                # Determinar usuario asignado
                usuario_asignado = None
                if sub_movil and sub_movil in usuario_ids:
                    usuario_asignado_id = usuario_ids[sub_movil]
                    usuario_asignado = User.objects.get(id=usuario_asignado_id)
                    no_asignados_flag = False
                else:
                    no_asignados += 1
                    no_asignados_flag = True

                # Crear cliente
                cliente = Cliente(
                    rut=rut_str,
                    dv=dv,
                    nombre_empresa=razon_social[:100] if razon_social else "Sin nombre",
                    sector="Telecom",
                    tipo_cliente="Empresa",
                    usuario_asignado=usuario_asignado,
                    estado="ACTIVO",
                    segmento=segmento[:50] if segmento else None,
                    subrango=subrango[:50] if subrango else None,
                    sub_movil=sub_movil[:50] if sub_movil else None,
                    subgerencia=subgerencia[:50] if subgerencia else None,
                    renta_uf_total=renta_uf,
                    q_total_lineas=q_total_lineas,
                    bam=bam,
                    m2m=m2m,
                    voz=voz,
                    fijo=fijo,
                    movil=movil,
                )
                cliente.save()
                importados += 1

                # Mostrar progreso cada 100 registros
                if importados % 100 == 0:
                    self.stdout.write(f"  {importados} clientes importados...")

            except Exception as e:
                errores += 1
                if errores <= 5:  # Mostrar solo los primeros 5 errores
                    self.stdout.write(
                        self.style.WARNING(f"⚠ Error en fila {row}: {str(e)}")
                    )

        # Resumen
        self.stdout.write("\n" + "="*80)
        self.stdout.write(self.style.SUCCESS(f"✓ IMPORTACIÓN COMPLETADA"))
        self.stdout.write("="*80)
        self.stdout.write(f"\n✓ Clientes importados: {importados}")
        self.stdout.write(f"⚠ Clientes duplicados: {duplicados}")
        self.stdout.write(f"⚠ Clientes sin asignar: {no_asignados}")
        self.stdout.write(f"✗ Errores: {errores}")
        self.stdout.write(f"\nTotal en BD: {Cliente.objects.count()}")
        self.stdout.write("\n" + "="*80)

        # Resumen por manager
        self.stdout.write(self.style.SUCCESS("\n📊 DISTRIBUCIÓN POR MANAGER:\n"))
        for vendedor, email in vendedor_a_usuario.items():
            try:
                user = User.objects.get(email=email)
                count = Cliente.objects.filter(usuario_asignado=user).count()
                self.stdout.write(f"  {vendedor:20s} ({email:30s}) → {count:5d} clientes")
            except:
                pass
