#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Management command para importar clientes desde Excel
Uso: python manage.py import_cartera_roy <ruta_archivo_excel>
"""

from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model
from core.models import Cliente
from pathlib import Path
import openpyxl
from decimal import Decimal

User = get_user_model()


class Command(BaseCommand):
    help = 'Importa clientes desde archivo Excel de cartera (Cartera de clientes Roy.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'archivo',
            type=str,
            help='Ruta al archivo Excel con la cartera de clientes'
        )
        parser.add_argument(
            '--usuario',
            type=str,
            help='Email del usuario ejecutivo a asignar (ej: roy@example.com)',
            default=None
        )

    def handle(self, *args, **options):
        archivo = Path(options['archivo'])
        usuario_email = options.get('usuario')
        
        # Validar que el archivo existe
        if not archivo.exists():
            raise CommandError(f'❌ Archivo no encontrado: {archivo}')
        
        self.stdout.write(self.style.SUCCESS(f'📂 Leyendo archivo: {archivo.name}'))
        
        try:
            wb = openpyxl.load_workbook(str(archivo))
            ws = wb.active
            
            # Obtener encabezados
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip().lower()] = col_idx
            
            self.stdout.write(f"📋 Encontrados {len(headers)} campos")
            
            # Buscar usuario para asignar
            usuario_asignado = None
            if usuario_email:
                try:
                    usuario_asignado = User.objects.get(email=usuario_email)
                    self.stdout.write(self.style.SUCCESS(f'✅ Usuario encontrado: {usuario_asignado.email}'))
                except User.DoesNotExist:
                    self.stdout.write(self.style.WARNING(f'⚠️  Usuario no encontrado: {usuario_email}'))
            
            # Contar filas a importar
            total_filas = ws.max_row - 1
            self.stdout.write(f"\n📊 Total de clientes a importar: {total_filas}")
            
            # Variables de control
            creados = 0
            actualizados = 0
            errores = 0
            
            # Iterar filas
            for row_idx in range(2, ws.max_row + 1):
                try:
                    # Extraer datos
                    rut_raw = ws.cell(row=row_idx, column=headers.get('rut', 1)).value
                    dv = ws.cell(row=row_idx, column=headers.get('dv', 2)).value
                    razon_social = ws.cell(row=row_idx, column=headers.get('razon_social', 3)).value
                    segmento = ws.cell(row=row_idx, column=headers.get('segmento', 4)).value
                    subrango = ws.cell(row=row_idx, column=headers.get('subrango', 5)).value
                    sub_movil = ws.cell(row=row_idx, column=headers.get('sub móvil', 6)).value
                    subgerencia = ws.cell(row=row_idx, column=headers.get('subgerencia', 7)).value
                    renta_uf = ws.cell(row=row_idx, column=headers.get('renta uf total', 8)).value
                    q_total_lineas = ws.cell(row=row_idx, column=headers.get('q total lineas', 9)).value
                    bam = ws.cell(row=row_idx, column=headers.get('bam', 10)).value
                    m2m = ws.cell(row=row_idx, column=headers.get('m2m', 11)).value
                    voz = ws.cell(row=row_idx, column=headers.get('voz', 12)).value
                    fijo = ws.cell(row=row_idx, column=headers.get('fijo', 13)).value
                    movil = ws.cell(row=row_idx, column=headers.get('móvil', 14)).value
                    fijo_movil = ws.cell(row=row_idx, column=headers.get('fijo+móvil', 15)).value
                    edv = ws.cell(row=row_idx, column=headers.get('edv', 16)).value
                    
                    # Validar datos requeridos
                    if not rut_raw or not razon_social:
                        errores += 1
                        continue
                    
                    # Limpiar RUT
                    rut_str = str(int(rut_raw)) if rut_raw else None
                    rut_limpio = f"{rut_str}-{dv}" if rut_str and dv else rut_str
                    
                    # Convertir valores numéricos
                    try:
                        renta_uf_decimal = Decimal(str(renta_uf)) if renta_uf else None
                    except:
                        renta_uf_decimal = None
                    
                    q_total = int(q_total_lineas) if q_total_lineas else 0
                    bam_val = int(bam) if bam else 0
                    m2m_val = int(m2m) if m2m else 0
                    voz_val = int(voz) if voz else 0
                    
                    # Crear o actualizar cliente
                    cliente, creado = Cliente.objects.update_or_create(
                        rut=rut_limpio,
                        defaults={
                            'dv': str(dv) if dv else None,
                            'nombre_empresa': str(razon_social).strip(),
                            'segmento': str(segmento).strip() if segmento else None,
                            'subrango': str(subrango).strip() if subrango else None,
                            'sub_movil': str(sub_movil).strip() if sub_movil else None,
                            'subgerencia': str(subgerencia).strip() if subgerencia else None,
                            'renta_uf_total': renta_uf_decimal,
                            'q_total_lineas': q_total,
                            'bam': bam_val,
                            'm2m': m2m_val,
                            'voz': voz_val,
                            'fijo': str(fijo).strip() if fijo else None,
                            'movil': str(movil).strip() if movil else None,
                            'fijo_movil': str(fijo_movil).strip() if fijo_movil else None,
                            'edv': str(edv).strip() if edv else None,
                            'usuario_asignado': usuario_asignado,
                            'estado': 'ACTIVO',
                        }
                    )
                    
                    if creado:
                        creados += 1
                    else:
                        actualizados += 1
                    
                    # Progreso
                    if (row_idx - 1) % 100 == 0:
                        self.stdout.write(f"  Procesadas {row_idx - 1} filas...")
                
                except Exception as e:
                    errores += 1
                    self.stdout.write(self.style.WARNING(f"⚠️  Error fila {row_idx}: {str(e)}"))
            
            # Resumen
            self.stdout.write("\n" + "=" * 80)
            self.stdout.write(self.style.SUCCESS(f"✅ IMPORTACIÓN COMPLETADA"))
            self.stdout.write("=" * 80)
            self.stdout.write(f"  📊 Total importado: {creados + actualizados}")
            self.stdout.write(f"  ✨ Clientes creados: {creados}")
            self.stdout.write(f"  🔄 Clientes actualizados: {actualizados}")
            self.stdout.write(f"  ❌ Errores: {errores}")
            self.stdout.write("=" * 80)
            
        except Exception as e:
            raise CommandError(f'❌ Error al procesar archivo: {str(e)}')
