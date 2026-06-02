#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Management command para importar clientes desde Excel - versión simplificada
Usa Raw SQL para evitar problemas de caché de Django
"""

from django.core.management.base import BaseCommand, CommandError
from django.db import connection
from pathlib import Path
import openpyxl
from decimal import Decimal
from django.utils import timezone


class Command(BaseCommand):
    help = 'Importa clientes desde archivo Excel de cartera'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo Excel')

    def handle(self, *args, **options):
        archivo = Path(options['archivo'])
        
        if not archivo.exists():
            raise CommandError(f'❌ Archivo no encontrado: {archivo}')
        
        self.stdout.write(self.style.SUCCESS(f'📂 Leyendo archivo: {archivo.name}'))
        
        try:
            wb = openpyxl.load_workbook(str(archivo))
            ws = wb.active
            
            # Obtener encabezados
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip().lower()] = col_idx
            
            self.stdout.write(f"📋 Encontrados {len(headers)} campos")
            
            # Contar filas a importar
            total_filas = ws.max_row - 1
            self.stdout.write(f"\n📊 Total de clientes a importar: {total_filas}")
            
            # Variables de control
            creados = 0
            actualizados = 0
            errores = 0
            
            # Usar conexión directa a BD
            cursor = connection.cursor()
            
            def safe_int(val):
                """Convierte valores numéricos, manejando 'NO', 'SI', etc."""
                if not val:
                    return 0
                val_str = str(val).strip().upper()
                if val_str in ('NO', 'SI', 'N/A', ''):
                    return 0
                try:
                    return int(float(str(val).replace(',', '.')))
                except:
                    return 0
            
            # Iterar filas
            for row_idx in range(2, ws.max_row + 1):
                try:
                    # Extraer datos
                    rut_raw = ws.cell(row=row_idx, column=headers.get('rut', 1)).value
                    dv = ws.cell(row=row_idx, column=headers.get('dv', 2)).value
                    razon_social = ws.cell(row=row_idx, column=headers.get('razon_social', 3)).value
                    segmento = ws.cell(row=row_idx, column=headers.get('segmento', 4)).value
                    subrango = ws.cell(row=row_idx, column=headers.get('subrango', 5)).value
                    supervisor = ws.cell(row=row_idx, column=headers.get('sub móvil', 6)).value
                    subgerencia = ws.cell(row=row_idx, column=headers.get('subgerencia', 7)).value
                    renta_uf = ws.cell(row=row_idx, column=headers.get('renta uf total', 8)).value
                    q_total_lineas = ws.cell(row=row_idx, column=headers.get('q total lineas', 9)).value
                    bam = ws.cell(row=row_idx, column=headers.get('bam', 10)).value
                    m2m = ws.cell(row=row_idx, column=headers.get('m2m', 11)).value
                    voz = ws.cell(row=row_idx, column=headers.get('voz', 12)).value
                    fijo = ws.cell(row=row_idx, column=headers.get('fijo', 13)).value
                    movil = ws.cell(row=row_idx, column=headers.get('móvil', 14)).value
                    fijo_movil = ws.cell(row=row_idx, column=headers.get('fijo+móvil', 15)).value
                    edv = ws.cell(row=row_idx, column=headers.get('edv', 16)).value
                    
                    # Validar datos requeridos
                    if not rut_raw or not razon_social:
                        errores += 1
                        continue
                    
                    # Limpiar RUT
                    rut_str = str(int(rut_raw)) if rut_raw else None
                    rut_limpio = f"{rut_str}-{dv}" if rut_str and dv else rut_str
                    
                    # Convertir valores numéricos (manejar "NO", "SI", etc.)
                    try:
                        renta_uf_decimal = float(renta_uf) if renta_uf and str(renta_uf).replace('.','').replace(',','').isdigit() else None
                    except:
                        renta_uf_decimal = None
                    
                    q_total = safe_int(q_total_lineas)
                    bam_val = safe_int(bam)
                    m2m_val = safe_int(m2m)
                    voz_val = safe_int(voz)
                    fijo_val = safe_int(fijo)
                    movil_val = safe_int(movil)
                    fijo_movil_val = safe_int(fijo_movil)
                    
                    # INSERT OR UPDATE usando SQL directo
                    now = timezone.now().isoformat()
                    sql = """
                    INSERT INTO core_cliente (
                        rut, nombre_empresa, estado, usuario_asignado_id, 
                        created_at, updated_at, dv, segmento, subrango, supervisor, 
                        subgerencia, renta_uf_total, q_total_lineas, bam, m2m, 
                        voz, fijo, movil, fijo_movil, edv
                    ) VALUES (%s, %s, %s, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (rut) DO UPDATE SET
                        nombre_empresa = EXCLUDED.nombre_empresa,
                        updated_at = EXCLUDED.updated_at,
                        dv = EXCLUDED.dv,
                        segmento = EXCLUDED.segmento,
                        subrango = EXCLUDED.subrango,
                        supervisor = EXCLUDED.supervisor,
                        subgerencia = EXCLUDED.subgerencia,
                        renta_uf_total = EXCLUDED.renta_uf_total,
                        q_total_lineas = EXCLUDED.q_total_lineas,
                        bam = EXCLUDED.bam,
                        m2m = EXCLUDED.m2m,
                        voz = EXCLUDED.voz,
                        fijo = EXCLUDED.fijo,
                        movil = EXCLUDED.movil,
                        fijo_movil = EXCLUDED.fijo_movil,
                        edv = EXCLUDED.edv
                    """
                    
                    cursor.execute(sql, [
                        rut_limpio, str(razon_social).strip(), 'ACTIVO', now, now,
                        str(dv)[:1] if dv else None,
                        str(segmento).strip() if segmento else None,
                        str(subrango).strip() if subrango else None,
                        str(supervisor).strip() if supervisor else None,
                        str(subgerencia).strip() if subgerencia else None,
                        renta_uf_decimal,
                        q_total, bam_val, m2m_val, voz_val, fijo_val, movil_val, fijo_movil_val,
                        str(edv).strip() if edv else None
                    ])
                    
                    if cursor.rowcount > 0:
                        creados += 1
                    else:
                        actualizados += 1
                    
                    # Progreso
                    if (row_idx - 1) % 100 == 0:
                        self.stdout.write(f"  Procesadas {row_idx - 1} filas...")
                
                except Exception as e:
                    errores += 1
                    if errores <= 5:  # Solo mostrar primeros 5 errores
                        self.stdout.write(self.style.WARNING(f"⚠️  Error fila {row_idx}: {str(e)}")[:100])
            
            connection.commit()
            cursor.close()
            
            # Resultado final
            self.stdout.write("\n" + "="*50)
            self.stdout.write(self.style.SUCCESS("✅ IMPORTACIÓN COMPLETADA"))
            self.stdout.write("="*50)
            self.stdout.write(f"📊 Total importado: {creados + actualizados}")
            self.stdout.write(self.style.SUCCESS(f"✅ Clientes creados: {creados}"))
            self.stdout.write(self.style.SUCCESS(f"🔄 Clientes actualizados: {actualizados}"))
            self.stdout.write(self.style.WARNING(f"⚠️  Errores: {errores}"))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Error: {str(e)}'))
