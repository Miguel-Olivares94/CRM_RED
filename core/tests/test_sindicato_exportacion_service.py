from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase

from openpyxl import load_workbook

from core.models import (
    AuditoriaSindicato,
    ConsolidadoDetalleSindicato,
    ConsolidadoMensualSindicato,
    Empresa,
    SocioSindicato,
    TipoBeneficioSindicato,
)
from core.services.sindicato_exportacion import ConsolidadoExportacionError, exportar_consolidado_excel


User = get_user_model()


class SindicatoExportacionServiceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa A', tipo='CLIENTE')
        cls.empresa_b = Empresa.objects.create(nombre='Empresa B', tipo='CLIENTE')
        cls.user = User.objects.create_user(username='tes_a', email='tes_a@test.cl', password='x')

        cls.benef_telefonia = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='TEL',
            nombre='Telefonia',
            orden_export=20,
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_gas = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            orden_export=10,
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_omi = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='OMI',
            nombre='Clinica OMI',
            orden_export=15,
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )

        cls.socio_1 = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio Uno',
            site='Site 1',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        cls.socio_2 = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='11111111-1',
            nombre='Socio Dos',
            site='Site 2',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_LICENCIA,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )

    def _crear_consolidado_cerrado(self):
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-10',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=2,
            total_monto=32000,
        )
        ConsolidadoDetalleSindicato.objects.create(
            empresa=self.empresa_a,
            consolidado=cons,
            socio=self.socio_1,
            tipo_beneficio=self.benef_gas,
            monto_aprobado=10000,
        )
        ConsolidadoDetalleSindicato.objects.create(
            empresa=self.empresa_a,
            consolidado=cons,
            socio=self.socio_1,
            tipo_beneficio=self.benef_omi,
            monto_aprobado=5000,
        )
        ConsolidadoDetalleSindicato.objects.create(
            empresa=self.empresa_a,
            consolidado=cons,
            socio=self.socio_2,
            tipo_beneficio=self.benef_telefonia,
            monto_aprobado=17000,
        )
        return cons

    def test_exporta_excel_de_consolidado_cerrado(self):
        cons = self._crear_consolidado_cerrado()
        result = exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons.id, usuario=self.user)
        self.assertTrue(result.filename.endswith('.xlsx'))
        self.assertGreater(len(result.content), 0)

    def test_bloquea_exportacion_si_esta_abierto(self):
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-11',
            estado=ConsolidadoMensualSindicato.ESTADO_ABIERTO,
            total_socios=0,
            total_monto=0,
        )
        with self.assertRaises(ConsolidadoExportacionError):
            exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons.id, usuario=self.user)

    def test_no_exporta_consolidado_de_otro_tenant(self):
        cons_b = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_b,
            periodo='2026-10',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=0,
            total_monto=0,
        )
        with self.assertRaises(ConsolidadoExportacionError):
            exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons_b.id, usuario=self.user)

    def test_columnas_respetan_orden_export(self):
        cons = self._crear_consolidado_cerrado()
        result = exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons.id, usuario=self.user)

        wb = load_workbook(BytesIO(result.content))
        ws = wb['Consolidado']
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[:4], ['RUT', 'Nombre', 'Site', 'Estado laboral'])
        self.assertEqual(headers[4:7], ['Gas', 'Clinica OMI', 'Telefonia'])
        self.assertEqual(headers[-1], 'Total General')

    def test_totales_por_socio_correctos(self):
        cons = self._crear_consolidado_cerrado()
        result = exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons.id, usuario=self.user)

        wb = load_workbook(BytesIO(result.content))
        ws = wb['Consolidado']
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Socio Dos primero por RUT, luego Socio Uno, luego fila TOTAL
        self.assertEqual(rows[0][0], '11111111-1')
        self.assertEqual(rows[0][-1], 17000)
        self.assertEqual(rows[1][0], '12345678-5')
        self.assertEqual(rows[1][-1], 15000)

        total_row = rows[-1]
        self.assertEqual(total_row[0], 'TOTAL')
        self.assertEqual(total_row[-1], 32000)

    def test_auditoria_creada_al_exportar(self):
        cons = self._crear_consolidado_cerrado()
        exportar_consolidado_excel(empresa=self.empresa_a, consolidado_id=cons.id, usuario=self.user)

        audit = AuditoriaSindicato.objects.filter(
            empresa=self.empresa_a,
            accion='EXPORTAR_CONSOLIDADO',
            periodo='2026-10',
        ).latest('created_at')
        self.assertEqual(audit.entidad, 'ConsolidadoMensualSindicato')
        self.assertEqual(audit.usuario, self.user)
        self.assertEqual(audit.payload.get('total_socios'), 2)
        self.assertEqual(audit.payload.get('total_monto'), 32000)
