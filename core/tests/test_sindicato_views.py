from io import BytesIO
from datetime import datetime

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from openpyxl import Workbook, load_workbook

import tempfile
import shutil

from django.test import override_settings

from core.models import (
    AlertaSindicato,
    AuditoriaSindicato,
    ConsolidadoMensualSindicato,
    DocumentoSindicato,
    Empresa,
    MovimientoSindicato,
    SocioSindicato,
    TipoBeneficioSindicato,
    UserProfile,
)


User = get_user_model()


class SindicatoViewsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa A', tipo='CLIENTE')
        cls.empresa_b = Empresa.objects.create(nombre='Empresa B', tipo='CLIENTE')

        cls.grp_admin = Group.objects.create(name='Administracion')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')
        cls.grp_dirigente = Group.objects.create(name='Dirigente')

        cls.admin_a = User.objects.create_user(username='admin_a', email='admin_a@test.cl', password='x')
        cls.admin_a.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_a, empresa=cls.empresa_a, role='ADMIN')

        cls.tesoreria_a = User.objects.create_user(username='tes_a', email='tes_a@test.cl', password='x')
        cls.tesoreria_a.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria_a, empresa=cls.empresa_a, role='ADMIN')

        cls.dirigente_a = User.objects.create_user(username='dir_a', email='dir_a@test.cl', password='x')
        cls.dirigente_a.groups.add(cls.grp_dirigente)
        UserProfile.objects.create(user=cls.dirigente_a, empresa=cls.empresa_a, role='ADMIN')

        cls.admin_b = User.objects.create_user(username='admin_b', email='admin_b@test.cl', password='x')
        cls.admin_b.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_b, empresa=cls.empresa_b, role='ADMIN')

        cls.socio_a = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio A',
        )
        cls.socio_b = SocioSindicato.objects.create(
            empresa=cls.empresa_b,
            rut='11111111-1',
            nombre='Socio B',
        )

        cls.benef_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_b = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_b,
            codigo='TEL',
            nombre='Telefonia',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )

        MovimientoSindicato.objects.create(
            empresa=cls.empresa_a,
            socio=cls.socio_a,
            tipo_beneficio=cls.benef_a,
            periodo='2026-06',
            monto=10000,
            observacion='ok',
            referencia_externa='A-1',
        )
        MovimientoSindicato.objects.create(
            empresa=cls.empresa_b,
            socio=cls.socio_b,
            tipo_beneficio=cls.benef_b,
            periodo='2026-06',
            monto=20000,
            observacion='ok',
            referencia_externa='B-1',
        )

    def test_usuario_no_autenticado_redirige_login(self):
        resp = self.client.get(reverse('core:sindicato_socio_list'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse('core:login'), resp.url)

    def test_usuario_autenticado_ve_datos_su_empresa(self):
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindicato_socio_list'))
        self.assertEqual(resp.status_code, 200)
        socios = list(resp.context['socios'])
        self.assertEqual(socios, [self.socio_a])

    def test_empresa_a_no_puede_editar_socio_empresa_b(self):
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindicato_socio_update', kwargs={'pk': self.socio_b.pk}))
        self.assertEqual(resp.status_code, 404)

    def test_listados_no_mezclan_tenant(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_movimiento_list'))
        self.assertEqual(resp.status_code, 200)
        movs = list(resp.context['movimientos'])
        self.assertEqual(len(movs), 1)
        self.assertEqual(movs[0].empresa, self.empresa_a)

    def test_consulta_rut_no_devuelve_datos_otro_tenant(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consulta_rut'), {'rut': '11.111.111-1'})
        self.assertEqual(resp.status_code, 200)
        self.assertIsNone(resp.context['socio'])

    def test_consulta_rut_tabla_socios_y_boton_ver_ficha(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consulta_rut'))
        self.assertEqual(resp.status_code, 200)

        socios_tabla = list(resp.context['socios_tabla'])
        self.assertEqual(socios_tabla, [self.socio_a])

        contenido = resp.content.decode('utf-8')
        self.assertIn('Socios disponibles', contenido)
        self.assertIn('Ver ficha', contenido)
        self.assertIn('?rut=12345678-5', contenido)
        self.assertNotIn('11.111.111-1', contenido)

    def test_consulta_rut_muestra_detalle_fuente_gas_con_vale(self):
        periodo_actual = datetime.now().strftime('%Y-%m')
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo=periodo_actual,
            monto=22222,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='GAS-RUT-1',
            fuente=MovimientoSindicato.FUENTE_GAS,
            metadata_fuente={'source_columns': {'vale_de_gas': '45 KG', 'site': 'SITE RUT'}},
        )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consulta_rut'), {'rut': self.socio_a.rut})
        self.assertEqual(resp.status_code, 200)

        contenido = resp.content.decode('utf-8')
        self.assertIn('Detalle fuente Gas', contenido)
        self.assertIn('45 KG', contenido)

    def test_consulta_rut_muestra_detalle_fuente_telefonia(self):
        periodo_actual = datetime.now().strftime('%Y-%m')
        benef_tel = TipoBeneficioSindicato.objects.create(
            empresa=self.empresa_a,
            codigo='TELA',
            nombre='Telefonia A',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=benef_tel,
            periodo=periodo_actual,
            monto=45678,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='TEL-RUT-1',
            fuente=MovimientoSindicato.FUENTE_TELEFONIA,
            metadata_fuente={
                'source_columns': {
                    'rut': '17.983.258-5',
                    'razon_social': 'SOCIO TEL TEST',
                    'cuenta': '99887766',
                    'pcs': '2',
                    'fecha_entrega': '2026-03-01',
                    'cargo_fijo': '45678',
                }
            },
        )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consulta_rut'), {'rut': self.socio_a.rut})
        self.assertEqual(resp.status_code, 200)

        contenido = resp.content.decode('utf-8')
        self.assertIn('Detalle fuente Telefonía', contenido)
        self.assertIn('99887766', contenido)
        self.assertIn('SOCIO TEL TEST', contenido)
        self.assertIn('45678', contenido)

    def test_consulta_rut_no_recorta_detalle_fuente_telefonia(self):
        periodo_actual = datetime.now().strftime('%Y-%m')
        benef_tel = TipoBeneficioSindicato.objects.create(
            empresa=self.empresa_a,
            codigo='TELB',
            nombre='Telefonia B',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )

        for idx in range(25):
            MovimientoSindicato.objects.create(
                empresa=self.empresa_a,
                socio=self.socio_a,
                tipo_beneficio=benef_tel,
                periodo=periodo_actual,
                monto=1000 + idx,
                estado=MovimientoSindicato.ESTADO_VALIDADO,
                referencia_externa=f'TEL-RUT-{idx}',
                fuente=MovimientoSindicato.FUENTE_TELEFONIA,
                metadata_fuente={'source_columns': {'cuenta': f'9000{idx}', 'pcs': str(idx), 'fecha_entrega': '2026-03-01'}},
            )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consulta_rut'), {'rut': self.socio_a.rut})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context['telefonia_movimientos_detalle']), 25)

    def test_socio_list_muestra_ficha_rapida_por_rut(self):
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindicato_socio_list'), {'rut': '12.345.678-5'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['socio_detalle'], self.socio_a)

        contenido = resp.content.decode('utf-8')
        self.assertIn('Ficha rápida del socio', contenido)
        self.assertIn('Editar perfil', contenido)

    def test_socio_list_muestra_detalle_fuente_gas_con_vale(self):
        periodo_actual = datetime.now().strftime('%Y-%m')
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo=periodo_actual,
            monto=12345,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='GAS-DET-1',
            fuente=MovimientoSindicato.FUENTE_GAS,
            metadata_fuente={'source_columns': {'vale_de_gas': '15 KG', 'site': 'SITE A'}},
        )

        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindicato_socio_list'), {'rut': self.socio_a.rut})
        self.assertEqual(resp.status_code, 200)

        contenido = resp.content.decode('utf-8')
        self.assertIn('Detalle fuente Gas', contenido)
        self.assertIn('15 KG', contenido)

    def test_socio_list_muestra_detalle_fuente_telefonia(self):
        periodo_actual = datetime.now().strftime('%Y-%m')
        benef_tel = TipoBeneficioSindicato.objects.create(
            empresa=self.empresa_a,
            codigo='TELS',
            nombre='Telefonia Socio',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=benef_tel,
            periodo=periodo_actual,
            monto=33333,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='TEL-SOC-1',
            fuente=MovimientoSindicato.FUENTE_TELEFONIA,
            metadata_fuente={'source_columns': {'cuenta': '11223344', 'pcs': '1', 'fecha_entrega': '2026-02-15'}},
        )

        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindicato_socio_list'), {'rut': self.socio_a.rut})
        self.assertEqual(resp.status_code, 200)

        contenido = resp.content.decode('utf-8')
        self.assertIn('Detalle fuente Telefonía', contenido)
        self.assertIn('11223344', contenido)

    def test_permisos_admin_crud_socios_beneficios(self):
        self.client.force_login(self.admin_a)
        resp_list = self.client.get(reverse('core:sindicato_socio_list'))
        resp_new = self.client.get(reverse('core:sindicato_socio_create'))
        resp_benef = self.client.get(reverse('core:sindicato_beneficio_create'))
        self.assertEqual(resp_list.status_code, 200)
        self.assertEqual(resp_new.status_code, 200)
        self.assertEqual(resp_benef.status_code, 200)

    def test_permisos_tesoreria_crud_movimientos_y_consulta(self):
        self.client.force_login(self.tesoreria_a)
        resp_mov = self.client.get(reverse('core:sindicato_movimiento_create'))
        resp_cons = self.client.get(reverse('core:sindicato_consulta_rut'))
        resp_socio_create = self.client.get(reverse('core:sindicato_socio_create'))
        self.assertEqual(resp_mov.status_code, 200)
        self.assertEqual(resp_cons.status_code, 200)
        self.assertEqual(resp_socio_create.status_code, 403)

    def test_permisos_dirigente_solo_lectura_y_consulta(self):
        self.client.force_login(self.dirigente_a)
        resp_socio_list = self.client.get(reverse('core:sindicato_socio_list'))
        resp_mov_list = self.client.get(reverse('core:sindicato_movimiento_list'))
        resp_cons = self.client.get(reverse('core:sindicato_consulta_rut'))
        resp_mov_create = self.client.get(reverse('core:sindicato_movimiento_create'))
        self.assertEqual(resp_socio_list.status_code, 200)
        self.assertEqual(resp_mov_list.status_code, 200)
        self.assertEqual(resp_cons.status_code, 200)
        self.assertEqual(resp_mov_create.status_code, 403)

    def test_create_movimiento_usa_empresa_usuario(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(
            reverse('core:sindicato_movimiento_create'),
            data={
                'socio': self.socio_a.pk,
                'tipo_beneficio': self.benef_a.pk,
                'periodo': '2026-07',
                'monto': '15000',
                'estado': 'PENDIENTE',
                'observacion': 'nuevo',
                'referencia_externa': 'A-2',
                'empresa_id': self.empresa_b.pk,  # intento de inyección
            },
        )
        self.assertEqual(resp.status_code, 302)
        mov = MovimientoSindicato.objects.get(referencia_externa='A-2')
        self.assertEqual(mov.empresa, self.empresa_a)

    def test_update_movimiento_no_permite_otro_tenant(self):
        mov_b = MovimientoSindicato.objects.get(empresa=self.empresa_b)
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_movimiento_update', kwargs={'pk': mov_b.pk}))
        self.assertEqual(resp.status_code, 404)


class SindicatoImportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa A', tipo='CLIENTE')
        cls.empresa_b = Empresa.objects.create(nombre='Empresa B', tipo='CLIENTE')

        cls.grp_admin = Group.objects.create(name='Administracion')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')
        cls.grp_dirigente = Group.objects.create(name='Dirigente')

        cls.admin_a = User.objects.create_user(username='admin_a', email='admin_a@test.cl', password='x')
        cls.admin_a.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_a, empresa=cls.empresa_a, role='ADMIN')

        cls.tesoreria_a = User.objects.create_user(username='tes_a', email='tes_a@test.cl', password='x')
        cls.tesoreria_a.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria_a, empresa=cls.empresa_a, role='ADMIN')

        cls.dirigente_a = User.objects.create_user(username='dir_a', email='dir_a@test.cl', password='x')
        cls.dirigente_a.groups.add(cls.grp_dirigente)
        UserProfile.objects.create(user=cls.dirigente_a, empresa=cls.empresa_a, role='ADMIN')

        cls.socio_a = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio A',
        )
        cls.socio_b = SocioSindicato.objects.create(
            empresa=cls.empresa_b,
            rut='11111111-1',
            nombre='Socio B',
        )

        cls.benef_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_tel_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='TEL',
            nombre='Telefonia',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_copeuch_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='COPA',
            nombre='Copeuch Activo',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_inactivo_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='COP',
            nombre='Copeuch',
            estado=TipoBeneficioSindicato.ESTADO_INACTIVO,
        )

    def _csv_file(self, content, name='movimientos.csv'):
        return SimpleUploadedFile(name, content.encode('utf-8'), content_type='text/csv')

    def _xlsx_file(self, rows, name='movimientos.xlsx'):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        payload = BytesIO()
        wb.save(payload)
        payload.seek(0)
        return SimpleUploadedFile(
            name,
            payload.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_usuario_sin_permiso_recibe_403(self):
        self.client.force_login(self.dirigente_a)
        resp = self.client.get(reverse('core:sindicato_movimiento_import'))
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(AuditoriaSindicato.objects.count(), 0)

    def test_archivo_valido_crea_movimientos(self):
        self.client.force_login(self.tesoreria_a)
        content = (
            'RUT,Nombre,Monto,Observacion,Referencia externa,Site\n'
            '12.345.678-5,Socio A,10000,ok,REF-OK-1,Site A\n'
            '11.111.111-1,Socio Nuevo Tenant A,12000,alta,REF-OK-2,Site B\n'
        )
        file_obj = self._csv_file(content)

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={
                'tipo_beneficio': self.benef_a.id,
                'periodo': '2026-06',
                'archivo': file_obj,
            },
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Válidas: 2')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 2')

        self.assertEqual(
            MovimientoSindicato.objects.filter(empresa=self.empresa_a, periodo='2026-06').count(),
            2,
        )
        socio_nuevo = SocioSindicato.objects.get(empresa=self.empresa_a, rut='11111111-1')
        self.assertEqual(socio_nuevo.nombre, 'Socio Nuevo Tenant A')
        self.assertEqual(SocioSindicato.objects.filter(empresa=self.empresa_b, rut='11111111-1').count(), 1)

    def test_rut_invalido_rechaza_fila(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto\n12.345.678-9,Socio Malo,10000\n'
        file_obj = self._csv_file(content)

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Rechazadas: 1')
        self.assertContains(preview, 'RUT inválido o vacío.')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 0')

    def test_monto_cero_rechaza_fila(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto\n12.345.678-5,Socio A,0\n'
        file_obj = self._csv_file(content)

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Monto debe ser mayor a cero.')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 0')

    def test_beneficio_inactivo_rechaza_importacion(self):
        self.client.force_login(self.admin_a)
        content = 'RUT,Nombre,Monto\n12.345.678-5,Socio A,10000\n'
        file_obj = self._csv_file(content)

        resp = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={
                'tipo_beneficio': self.benef_inactivo_a.id,
                'periodo': '2026-06',
                'archivo': file_obj,
            },
        )
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'está inactivo')
        self.assertEqual(MovimientoSindicato.objects.filter(empresa=self.empresa_a).count(), 0)

    def test_duplicado_con_referencia_externa_rechaza(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-06',
            monto=10000,
            referencia_externa='REF-DUP',
        )
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto,Referencia externa\n12.345.678-5,Socio A,10000,REF-DUP\n'
        file_obj = self._csv_file(content)

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Referencia externa duplicada')
        self.assertContains(preview, 'Rechazadas: 1')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 0')

    def test_tenant_isolation_importa_solo_empresa_usuario(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto,Referencia externa\n11.111.111-1,Socio Tenant A,9000,REF-TENANT-1\n'
        file_obj = self._csv_file(content)

        self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-07', 'archivo': file_obj},
        )
        self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})

        mov = MovimientoSindicato.objects.get(referencia_externa='REF-TENANT-1')
        self.assertEqual(mov.empresa, self.empresa_a)
        self.assertEqual(mov.socio.empresa, self.empresa_a)
        self.assertEqual(SocioSindicato.objects.filter(empresa=self.empresa_b, rut='11111111-1').count(), 1)

    def test_confirmar_importacion_crea_auditoria(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto,Referencia externa\n12.345.678-5,Socio A,10000,REF-AUD-1\n'
        file_obj = self._csv_file(content, name='import_auditoria.csv')

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)

        audit = AuditoriaSindicato.objects.latest('created_at')
        self.assertEqual(audit.accion, 'IMPORTAR_MOVIMIENTOS')
        self.assertEqual(audit.entidad, 'MovimientoSindicato')
        self.assertEqual(audit.periodo, '2026-06')
        self.assertEqual(audit.empresa, self.empresa_a)
        self.assertEqual(audit.usuario, self.tesoreria_a)
        self.assertEqual(audit.payload.get('total_filas_leidas'), 1)
        self.assertEqual(audit.payload.get('total_importadas'), 1)
        self.assertEqual(audit.payload.get('total_rechazadas'), 0)
        self.assertEqual(audit.payload.get('nombre_archivo'), 'import_auditoria.csv')

    def test_auditoria_queda_asociada_a_empresa_correcta(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto,Referencia externa\n11.111.111-1,Socio Tenant A,9000,REF-AUD-2\n'
        file_obj = self._csv_file(content, name='import_tenant.csv')

        self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-07', 'archivo': file_obj},
        )
        self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})

        audit = AuditoriaSindicato.objects.latest('created_at')
        self.assertEqual(audit.empresa, self.empresa_a)
        self.assertNotEqual(audit.empresa, self.empresa_b)

    def test_importacion_persiste_fuente_y_metadata(self):
        self.client.force_login(self.tesoreria_a)
        content = 'RUT,Nombre,Monto,Referencia externa\n12.345.678-5,Socio A,10000,REF-META-1\n'
        file_obj = self._csv_file(content, name='fuente_gas.csv')

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)

        mov = MovimientoSindicato.objects.get(referencia_externa='REF-META-1')
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_GAS)
        self.assertEqual(mov.metadata_fuente.get('file_name'), 'fuente_gas.csv')
        self.assertEqual(mov.metadata_fuente.get('source_row'), 2)
        self.assertEqual(mov.metadata_fuente.get('imported_by'), self.tesoreria_a.username)
        self.assertEqual(mov.metadata_fuente.get('source_columns', {}).get('monto'), '10000')

    def _importar_y_confirmar(self, *, content, beneficio_id, periodo='2026-06', name='import.csv'):
        file_obj = self._csv_file(content, name=name)
        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': beneficio_id, 'periodo': periodo, 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        return preview, confirm

    def test_copeuch_tot_dctos_con_puntos_crea_monto_correcto(self):
        self.client.force_login(self.tesoreria_a)
        content = (
            'RUT,NOMBRE,FEC. ING. SOCIO,ACCIONES,PRESTAMOS,TOT. DCTOS.\n'
            '12.345.678-5,Socio A,2020-01-01,0,0,3.680\n'
        )

        _preview, confirm = self._importar_y_confirmar(
            content=content,
            beneficio_id=self.benef_copeuch_a.id,
            name='copeuch_tot_dctos_puntos.csv',
        )
        self.assertContains(confirm, 'Movimientos creados: 1')

        mov = MovimientoSindicato.objects.get(empresa=self.empresa_a, referencia_externa__startswith='COP-')
        self.assertEqual(int(mov.monto), 3680)
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_COPEUCH)

    def test_copeuch_total_descuentos_crea_monto_correcto(self):
        self.client.force_login(self.tesoreria_a)
        content = (
            'RUT,NOMBRE,Total Descuentos\n'
            '12.345.678-5,Socio A,4200\n'
        )

        _preview, confirm = self._importar_y_confirmar(
            content=content,
            beneficio_id=self.benef_copeuch_a.id,
            name='copeuch_total_descuentos.csv',
        )
        self.assertContains(confirm, 'Movimientos creados: 1')

        mov = MovimientoSindicato.objects.get(empresa=self.empresa_a, referencia_externa__startswith='COP-')
        self.assertEqual(int(mov.monto), 4200)
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_COPEUCH)

    def test_copeuch_headers_con_acentos_puntos_y_espacios_crea_monto_correcto(self):
        self.client.force_login(self.tesoreria_a)
        content = (
            ' RUT , NOMBRE , FEC. ING. SOCIO , ACCIONES , PRÉSTAMOS , Total  Dctos  \n'
            '12.345.678-5,Socio A,2020-01-01,0,0,5100\n'
        )

        _preview, confirm = self._importar_y_confirmar(
            content=content,
            beneficio_id=self.benef_copeuch_a.id,
            name='copeuch_headers_ruidosos.csv',
        )
        self.assertContains(confirm, 'Movimientos creados: 1')

        mov = MovimientoSindicato.objects.get(empresa=self.empresa_a, referencia_externa__startswith='COP-')
        self.assertEqual(int(mov.monto), 5100)
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_COPEUCH)

    def test_copeuch_excel_detecta_header_despues_de_filas_informativas(self):
        self.client.force_login(self.tesoreria_a)
        file_obj = self._xlsx_file(
            [
                ('COOPEUCH LTDA.', None, None, None, None, None, None),
                ('Empleado:', 'SINDICATO DHL', None, None, None, None, None),
                ('Fecha', 'MAYO/2026', None, None, None, None, None),
                ('Planilla', '1315105', None, None, None, None, None),
                (None, None, None, None, None, None, None),
                ('NRO.', 'RUT', 'NOMBRE', 'FEC. ING. SOCIO', 'ACCIONES', 'PRESTAMOS', 'TOT. DCTOS.'),
                (1, '12.345.678-5', 'Socio A', '28-10-25', 3680, 0, 3680),
            ],
            name='copeuch_con_bloque_informativo.xlsx',
        )

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_copeuch_a.id, 'periodo': '2026-08', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Válidas: 1')
        self.assertContains(preview, 'Rechazadas: 0')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 1')

        mov = MovimientoSindicato.objects.get(empresa=self.empresa_a, periodo='2026-08')
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_COPEUCH)
        self.assertEqual(int(mov.monto), 3680)

    def test_gas_y_telefonia_siguen_detectandose_correctamente(self):
        self.client.force_login(self.tesoreria_a)
        gas = (
            'RUT,NOMBRE APELLIDO,SITE,VALE DE GAS,MONTO\n'
            '12.345.678-5,Socio A,Site A,15 KG,10000\n'
        )
        tel = (
            'RUT,Razon social,Cuenta,PCS,Cargo Fijo,Fecha de entrega\n'
            '12.345.678-5,Socio A,C1,1,7000,2026-06-01\n'
        )

        _pg, cg = self._importar_y_confirmar(content=gas, beneficio_id=self.benef_a.id, name='gas_ok.csv')
        self.assertContains(cg, 'Movimientos creados: 1')
        gas_mov = MovimientoSindicato.objects.filter(
            empresa=self.empresa_a,
            fuente=MovimientoSindicato.FUENTE_GAS,
        ).latest('id')
        self.assertEqual(gas_mov.fuente, MovimientoSindicato.FUENTE_GAS)

        _pt, ct = self._importar_y_confirmar(content=tel, beneficio_id=self.benef_tel_a.id, name='tel_ok.csv')
        self.assertContains(ct, 'Movimientos creados: 1')
        tel_mov = MovimientoSindicato.objects.filter(empresa=self.empresa_a, fuente=MovimientoSindicato.FUENTE_TELEFONIA).latest('id')
        self.assertEqual(int(tel_mov.monto), 7000)

    def test_preview_telefonia_muestra_campos_fuente(self):
        self.client.force_login(self.tesoreria_a)
        tel = (
            'RUT,Razon social,Cuenta,PCS,Cargo Fijo,Fecha de entrega\n'
            '12.345.678-5,Socio A,C1,1,7000,2026-06-01\n'
        )
        file_obj = self._csv_file(tel, name='tel_preview.csv')

        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={'tipo_beneficio': self.benef_tel_a.id, 'periodo': '2026-06', 'archivo': file_obj},
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Cuenta')
        self.assertContains(preview, 'Fecha entrega')
        self.assertContains(preview, 'C1')
        self.assertContains(preview, '2026-06-01')


class SindicatoConsolidadoViewsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa A', tipo='CLIENTE')
        cls.empresa_b = Empresa.objects.create(nombre='Empresa B', tipo='CLIENTE')

        cls.grp_admin = Group.objects.create(name='Administracion')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')
        cls.grp_dirigente = Group.objects.create(name='Dirigente')

        cls.admin_a = User.objects.create_user(username='admin_a', email='admin_a@test.cl', password='x')
        cls.admin_a.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_a, empresa=cls.empresa_a, role='ADMIN')

        cls.tesoreria_a = User.objects.create_user(username='tes_a', email='tes_a@test.cl', password='x')
        cls.tesoreria_a.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria_a, empresa=cls.empresa_a, role='ADMIN')

        cls.dirigente_a = User.objects.create_user(username='dir_a', email='dir_a@test.cl', password='x')
        cls.dirigente_a.groups.add(cls.grp_dirigente)
        UserProfile.objects.create(user=cls.dirigente_a, empresa=cls.empresa_a, role='ADMIN')

        cls.socio_a = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio A',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
            orden_export=1,
        )

        cls.socio_b = SocioSindicato.objects.create(
            empresa=cls.empresa_b,
            rut='11111111-1',
            nombre='Socio B',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_b = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_b,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
            orden_export=1,
        )

    def test_dirigente_ve_historial_pero_no_genera(self):
        self.client.force_login(self.dirigente_a)
        resp_hist = self.client.get(reverse('core:sindicato_consolidado_historial'))
        self.assertEqual(resp_hist.status_code, 200)

        resp_gen = self.client.post(
            reverse('core:sindicato_consolidado_generar'),
            data={'periodo': '2026-08'},
        )
        self.assertEqual(resp_gen.status_code, 403)

    def test_generar_consolidado_desde_view_usa_servicio(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-08',
            monto=10000,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='A-1',
        )
        self.client.force_login(self.tesoreria_a)

        resp = self.client.post(
            reverse('core:sindicato_consolidado_generar'),
            data={'periodo': '2026-08'},
        )
        self.assertEqual(resp.status_code, 302)

        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-08')
        self.assertEqual(cons.total_socios, 1)
        self.assertEqual(cons.total_monto, 10000)
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='GENERAR_CONSOLIDADO',
                periodo='2026-08',
            ).exists()
        )

    def test_cerrar_periodo_desde_view(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-09',
            monto=5000,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='A-2',
        )
        self.client.force_login(self.admin_a)
        self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': '2026-09'})

        resp_cerrar = self.client.post(
            reverse('core:sindicato_consolidado_cerrar'),
            data={'periodo': '2026-09'},
        )
        self.assertEqual(resp_cerrar.status_code, 302)

        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-09')
        self.assertEqual(cons.estado, ConsolidadoMensualSindicato.ESTADO_CERRADO)
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='CERRAR_CONSOLIDADO',
                periodo='2026-09',
            ).exists()
        )

    def test_historial_no_mezcla_tenant(self):
        ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_b,
            periodo='2026-08',
            estado=ConsolidadoMensualSindicato.ESTADO_ABIERTO,
            total_socios=1,
            total_monto=1000,
        )
        self.client.force_login(self.tesoreria_a)

        resp = self.client.get(reverse('core:sindicato_consolidado_historial'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(list(resp.context['consolidados']), [])

    def test_dirigente_recibe_403_en_exportacion(self):
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-12',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=0,
            total_monto=0,
        )
        self.client.force_login(self.dirigente_a)
        resp = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': cons.pk}))
        self.assertEqual(resp.status_code, 403)

    def test_no_exporta_consolidado_otro_tenant(self):
        cons_b = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_b,
            periodo='2026-12',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=0,
            total_monto=0,
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': cons_b.pk}))
        self.assertEqual(resp.status_code, 302)

    def test_exporta_excel_consolidado_cerrado_desde_endpoint(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-12',
            monto=8000,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='EXP-1',
        )
        self.client.force_login(self.admin_a)
        self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': '2026-12'})
        self.client.post(reverse('core:sindicato_consolidado_cerrar'), data={'periodo': '2026-12'})
        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-12')

        resp = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': cons.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resp['Content-Type'],
        )
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='EXPORTAR_CONSOLIDADO',
                periodo='2026-12',
            ).exists()
        )

    def test_generar_aplica_prevalidacion_licencia_y_baja(self):
        socio_lic = SocioSindicato.objects.create(
            empresa=self.empresa_a,
            rut='22222222-2',
            nombre='Socio Licencia',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        socio_baja = SocioSindicato.objects.create(
            empresa=self.empresa_a,
            rut='33333333-3',
            nombre='Socio Baja',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        mov_lic = MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=socio_lic,
            tipo_beneficio=self.benef_a,
            periodo='2026-10',
            monto=10000,
            estado=MovimientoSindicato.ESTADO_PENDIENTE,
            observacion='licencia medica',
            referencia_externa='PREV-LIC-1',
        )
        mov_baja = MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=socio_baja,
            tipo_beneficio=self.benef_a,
            periodo='2026-10',
            monto=12000,
            estado=MovimientoSindicato.ESTADO_PENDIENTE,
            observacion='baja',
            referencia_externa='PREV-BAJA-1',
        )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': '2026-10'})
        self.assertEqual(resp.status_code, 302)

        mov_lic.refresh_from_db()
        mov_baja.refresh_from_db()
        socio_lic.refresh_from_db()
        socio_baja.refresh_from_db()
        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-10')

        self.assertEqual(mov_lic.estado, MovimientoSindicato.ESTADO_OBSERVADO)
        self.assertEqual(socio_lic.estado_laboral, SocioSindicato.ESTADO_LABORAL_LICENCIA)
        self.assertEqual(mov_baja.estado, MovimientoSindicato.ESTADO_RECHAZADO)
        self.assertEqual(socio_baja.estado_laboral, SocioSindicato.ESTADO_LABORAL_DESVINCULADO)
        self.assertEqual(cons.total_socios, 1)
        self.assertEqual(cons.total_monto, 10000)
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='PREVALIDAR_CONSOLIDADO',
                periodo='2026-10',
            ).exists()
        )

    def test_detalle_consolidado_renderiza_sin_nameerror(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-10',
            monto=6000,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='DET-1',
        )
        self.client.force_login(self.tesoreria_a)
        self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': '2026-10'})
        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-10')

        resp = self.client.get(reverse('core:sindicato_consolidado_detalle', kwargs={'pk': cons.pk}))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(list(resp.context['detalles'])[0].monto_aprobado, 6000)

    def test_exportar_consolidado_ya_exportado_sigue_disponible(self):
        MovimientoSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_beneficio=self.benef_a,
            periodo='2026-11',
            monto=7000,
            estado=MovimientoSindicato.ESTADO_VALIDADO,
            referencia_externa='EXP-REPEAT-1',
        )
        self.client.force_login(self.admin_a)
        self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': '2026-11'})
        self.client.post(reverse('core:sindicato_consolidado_cerrar'), data={'periodo': '2026-11'})
        cons = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa_a, periodo='2026-11')

        first = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': cons.pk}))
        self.assertEqual(first.status_code, 200)
        cons.refresh_from_db()
        self.assertEqual(cons.estado, ConsolidadoMensualSindicato.ESTADO_EXPORTADO)

        second = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': cons.pk}))
        self.assertEqual(second.status_code, 200)


class SindiAppNavegacionTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa Nav', tipo='CLIENTE')

        cls.grp_admin = Group.objects.create(name='Administracion')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')
        cls.grp_dirigente = Group.objects.create(name='Dirigente')

        cls.admin_a = User.objects.create_user(username='admin_nav', email='admin_nav@test.cl', password='x')
        cls.admin_a.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_a, empresa=cls.empresa_a, role='ADMIN')

        cls.tesoreria_a = User.objects.create_user(username='tes_nav', email='tes_nav@test.cl', password='x')
        cls.tesoreria_a.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria_a, empresa=cls.empresa_a, role='ADMIN')

        cls.socio_a = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio Nav',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
            orden_export=1,
        )

    def test_sidebar_no_tiene_items_duplicados_ni_importar_ni_exportacion(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_dashboard'))
        self.assertEqual(resp.status_code, 200)
        contenido = resp.content.decode('utf-8')

        self.assertEqual(contenido.count('<li class="nav-item">'), 9)
        self.assertNotIn('Importar planillas</a></li>', contenido)
        self.assertNotIn('Exportación</a></li>', contenido)

    def test_rutas_importar_y_exportacion_siguen_funcionando_sin_estar_en_sidebar(self):
        self.client.force_login(self.tesoreria_a)
        resp_importar = self.client.get(reverse('core:sindiapp_movimiento_import'))
        resp_exportacion = self.client.get(reverse('core:sindiapp_exportacion_list'))
        self.assertEqual(resp_importar.status_code, 200)
        self.assertEqual(resp_exportacion.status_code, 200)

    def test_dashboard_movimientos_consolidados_alertas_navegables(self):
        self.client.force_login(self.tesoreria_a)
        for url_name in (
            'sindiapp_dashboard',
            'sindiapp_movimiento_list',
            'sindiapp_consolidado_historial',
            'sindiapp_alerta_list',
        ):
            resp = self.client.get(reverse(f'core:{url_name}'))
            self.assertEqual(resp.status_code, 200, url_name)

    def test_movimientos_muestra_boton_importar_planillas(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_movimiento_list'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Importar planillas')
        self.assertContains(resp, reverse('core:sindiapp_movimiento_import'))

    def test_dashboard_sin_alertas_criticas_no_muestra_cta(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_dashboard'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['alertas_criticas_count'], 0)
        self.assertNotContains(resp, 'alerta crítica pendiente')
        self.assertNotContains(resp, 'alertas críticas pendientes')

    def test_dashboard_muestra_cta_cuando_hay_alerta_critica(self):
        AlertaSindicato.objects.create(
            empresa=self.empresa_a,
            socio=self.socio_a,
            tipo_alerta='TEST',
            categoria=AlertaSindicato.CATEGORIA_DATOS,
            prioridad=AlertaSindicato.PRIORIDAD_CRITICA,
            titulo='Alerta crítica de prueba',
            estado=AlertaSindicato.ESTADO_PENDIENTE,
            clave_unica='ALERTA-CRITICA-NAV-1',
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_dashboard'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['alertas_criticas_count'], 1)
        self.assertContains(resp, 'alerta crítica pendiente')
        self.assertContains(resp, reverse('core:sindiapp_alerta_list'))

    def test_consolidado_historial_muestra_boton_exportar_para_periodo_cerrado(self):
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-09',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=1,
            total_monto=1000,
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_consolidado_historial'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Exportar')
        self.assertContains(resp, reverse('core:sindiapp_consolidado_exportar', kwargs={'pk': cons.pk}))

    def test_consolidado_historial_no_muestra_exportar_para_periodo_abierto(self):
        ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-10',
            estado=ConsolidadoMensualSindicato.ESTADO_ABIERTO,
            total_socios=1,
            total_monto=1000,
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_consolidado_historial'))
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, '>Exportar<')

    def test_auditoria_lista_registros_y_kpis(self):
        AuditoriaSindicato.objects.create(
            empresa=self.empresa_a,
            usuario=self.tesoreria_a,
            accion='IMPORTAR_MOVIMIENTOS',
            entidad='MovimientoSindicato',
            entidad_id='1',
            periodo='2026-06',
            resumen='Importación de prueba',
        )
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-06',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=1,
            total_monto=1000,
        )
        AuditoriaSindicato.objects.create(
            empresa=self.empresa_a,
            usuario=self.admin_a,
            accion='CERRAR_CONSOLIDADO',
            entidad='ConsolidadoMensualSindicato',
            entidad_id=str(cons.pk),
            periodo='2026-06',
            resumen='Cierre de prueba',
        )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_auditoria_list'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context['auditorias']), 2)
        self.assertContains(resp, 'Importación de movimientos')
        self.assertContains(resp, 'Cierre de período')
        self.assertContains(resp, reverse('core:sindiapp_consolidado_detalle', kwargs={'pk': cons.pk}))

    def test_auditoria_filtra_por_tipo_de_accion(self):
        AuditoriaSindicato.objects.create(
            empresa=self.empresa_a,
            usuario=self.tesoreria_a,
            accion='IMPORTAR_MOVIMIENTOS',
            entidad='MovimientoSindicato',
            entidad_id='1',
            periodo='2026-06',
            resumen='Importación de prueba',
        )
        AuditoriaSindicato.objects.create(
            empresa=self.empresa_a,
            usuario=self.admin_a,
            accion='EXPORTAR_CONSOLIDADO',
            entidad='ConsolidadoMensualSindicato',
            entidad_id='99',
            periodo='2026-06',
            resumen='Exportación de prueba',
        )

        self.client.force_login(self.tesoreria_a)
        resp = self.client.get(reverse('core:sindiapp_auditoria_list'), {'accion': 'EXPORTAR_CONSOLIDADO'})
        self.assertEqual(resp.status_code, 200)
        auditorias = list(resp.context['auditorias'])
        self.assertEqual(len(auditorias), 1)
        self.assertEqual(auditorias[0].accion, 'EXPORTAR_CONSOLIDADO')
        self.assertContains(resp, 'Exportación de prueba')
        self.assertNotContains(resp, 'Importación de prueba')

    def test_consolidado_historial_dirigente_no_ve_boton_exportar(self):
        cons = ConsolidadoMensualSindicato.objects.create(
            empresa=self.empresa_a,
            periodo='2026-11',
            estado=ConsolidadoMensualSindicato.ESTADO_CERRADO,
            total_socios=1,
            total_monto=1000,
        )
        dirigente_a = User.objects.create_user(username='dir_nav', email='dir_nav@test.cl', password='x')
        dirigente_a.groups.add(self.grp_dirigente)
        UserProfile.objects.create(user=dirigente_a, empresa=self.empresa_a, role='ADMIN')

        self.client.force_login(dirigente_a)
        resp = self.client.get(reverse('core:sindiapp_consolidado_historial'))
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, reverse('core:sindiapp_consolidado_exportar', kwargs={'pk': cons.pk}))


class SindicatoConsolidadoE2EFlowTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.empresa = Empresa.objects.create(nombre='Empresa E2E', tipo='CLIENTE')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')

        cls.tesoreria = User.objects.create_user(username='tes_e2e', email='tes_e2e@test.cl', password='x')
        cls.tesoreria.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria, empresa=cls.empresa, role='ADMIN')

        cls.benef = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
            orden_export=1,
        )

    def test_flujo_e2e_importar_generar_detalle_cerrar_exportar_y_validar_excel(self):
        self.client.force_login(self.tesoreria)
        periodo = '2026-12'

        # 1) Importar movimientos
        csv_content = (
            'RUT,Nombre,Monto,Observacion,Referencia externa,Site\n'
            '12.345.678-5,Socio Uno,10000,ok,REF-E2E-1,Site Norte\n'
            '11.111.111-1,Socio Dos,15000,ok,REF-E2E-2,Site Sur\n'
        )
        preview = self.client.post(
            reverse('core:sindicato_movimiento_import'),
            data={
                'tipo_beneficio': self.benef.id,
                'periodo': periodo,
                'archivo': SimpleUploadedFile('e2e_movs.csv', csv_content.encode('utf-8'), content_type='text/csv'),
            },
        )
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, 'Válidas: 2')

        confirm = self.client.post(reverse('core:sindicato_movimiento_import'), data={'action': 'confirmar'})
        self.assertEqual(confirm.status_code, 200)
        self.assertContains(confirm, 'Movimientos creados: 2')
        self.assertEqual(MovimientoSindicato.objects.filter(empresa=self.empresa, periodo=periodo).count(), 2)

        # 2) Generar consolidado
        generar = self.client.post(reverse('core:sindicato_consolidado_generar'), data={'periodo': periodo})
        self.assertEqual(generar.status_code, 302)
        consolidado = ConsolidadoMensualSindicato.objects.get(empresa=self.empresa, periodo=periodo)
        self.assertEqual(consolidado.estado, ConsolidadoMensualSindicato.ESTADO_ABIERTO)
        self.assertEqual(consolidado.total_socios, 2)
        self.assertEqual(int(consolidado.total_monto), 25000)

        # 3) Ver detalle
        detalle = self.client.get(reverse('core:sindicato_consolidado_detalle', kwargs={'pk': consolidado.pk}))
        self.assertEqual(detalle.status_code, 200)
        self.assertEqual(len(list(detalle.context['detalles'])), 2)

        # 4) Cerrar período
        cerrar = self.client.post(reverse('core:sindicato_consolidado_cerrar'), data={'periodo': periodo})
        self.assertEqual(cerrar.status_code, 302)
        consolidado.refresh_from_db()
        self.assertEqual(consolidado.estado, ConsolidadoMensualSindicato.ESTADO_CERRADO)

        # 5) Exportar Excel
        export = self.client.get(reverse('core:sindicato_consolidado_exportar', kwargs={'pk': consolidado.pk}))
        self.assertEqual(export.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export['Content-Type'],
        )

        # 6) Validar contenido del archivo generado
        wb = load_workbook(filename=BytesIO(export.content))
        ws = wb['Consolidado']
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[:4], ['RUT', 'Nombre', 'Site', 'Estado laboral'])
        self.assertIn('Gas', headers)
        self.assertEqual(headers[-1], 'Total General')

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total_row = None
        for row in rows:
            if row[0] == 'TOTAL':
                total_row = row
                break
        self.assertIsNotNone(total_row)

        gas_idx = headers.index('Gas')
        total_idx = len(headers) - 1
        sum_gas = sum(int(r[gas_idx] or 0) for r in rows if r[0] != 'TOTAL')
        sum_total = sum(int(r[total_idx] or 0) for r in rows if r[0] != 'TOTAL')
        self.assertEqual(int(total_row[gas_idx] or 0), sum_gas)
        self.assertEqual(int(total_row[total_idx] or 0), sum_total)


MEDIA_TEMP = tempfile.mkdtemp()


@override_settings(MEDIA_ROOT=MEDIA_TEMP)
class DocumentoSindicatoTests(TestCase):
    """Tests para el módulo de carga inteligente de documentos (OCR)."""

    @classmethod
    def setUpTestData(cls):
        cls.empresa_a = Empresa.objects.create(nombre='Empresa Doc A', tipo='CLIENTE')
        cls.empresa_b = Empresa.objects.create(nombre='Empresa Doc B', tipo='CLIENTE')

        cls.grp_admin = Group.objects.create(name='Administracion')
        cls.grp_tesoreria = Group.objects.create(name='Tesoreria')
        cls.grp_dirigente = Group.objects.create(name='Dirigente')

        cls.admin_a = User.objects.create_user(
            username='admin_doc_a', email='admin_doc_a@test.cl', password='x'
        )
        cls.admin_a.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_a, empresa=cls.empresa_a, role='ADMIN')

        cls.tesoreria_a = User.objects.create_user(
            username='tes_doc_a', email='tes_doc_a@test.cl', password='x'
        )
        cls.tesoreria_a.groups.add(cls.grp_tesoreria)
        UserProfile.objects.create(user=cls.tesoreria_a, empresa=cls.empresa_a, role='ADMIN')

        cls.dirigente_a = User.objects.create_user(
            username='dir_doc_a', email='dir_doc_a@test.cl', password='x'
        )
        cls.dirigente_a.groups.add(cls.grp_dirigente)
        UserProfile.objects.create(user=cls.dirigente_a, empresa=cls.empresa_a, role='ADMIN')

        cls.admin_b = User.objects.create_user(
            username='admin_doc_b', email='admin_doc_b@test.cl', password='x'
        )
        cls.admin_b.groups.add(cls.grp_admin)
        UserProfile.objects.create(user=cls.admin_b, empresa=cls.empresa_b, role='ADMIN')

        cls.socio_a = SocioSindicato.objects.create(
            empresa=cls.empresa_a,
            rut='12345678-5',
            nombre='Socio A',
            estado_laboral=SocioSindicato.ESTADO_LABORAL_ACTIVO,
            estado=SocioSindicato.ESTADO_ACTIVO,
        )
        cls.benef_a = TipoBeneficioSindicato.objects.create(
            empresa=cls.empresa_a,
            codigo='GAS',
            nombre='Gas',
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
            orden_export=1,
        )

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(MEDIA_TEMP, ignore_errors=True)
        super().tearDownClass()

    def _jpg_file(self, name='test.jpg'):
        from PIL import Image
        import io
        img = Image.new('RGB', (100, 50), color='white')
        buf = io.BytesIO()
        img.save(buf, format='JPEG')
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='image/jpeg')

    def _png_file(self, name='test.png'):
        from PIL import Image
        import io
        img = Image.new('RGB', (100, 50), color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='image/png')

    def _pdf_file(self, name='test.pdf'):
        content = b'%PDF-1.4\n1 0 obj\n<< /Type /Catalog >>\nendobj\nxref\n0 2\n%%EOF'
        return SimpleUploadedFile(name, content, content_type='application/pdf')

    # ---- Listado ----

    def test_listado_requiere_autenticacion(self):
        resp = self.client.get(reverse('core:sindiapp_documento_list'))
        self.assertEqual(resp.status_code, 302)

    def test_listado_accesible_para_admin_tesoreria_dirigente(self):
        for user in (self.admin_a, self.tesoreria_a, self.dirigente_a):
            self.client.force_login(user)
            resp = self.client.get(reverse('core:sindiapp_documento_list'))
            self.assertEqual(resp.status_code, 200, user.username)

    def test_listado_no_mezcla_tenant(self):
        DocumentoSindicato.objects.create(
            empresa=self.empresa_b,
            subido_por=self.admin_b,
            nombre_archivo='otro_tenant.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_SUBIDO,
        )
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindiapp_documento_list'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(list(resp.context['documentos']), [])

    # ---- Subir ----

    def test_subir_requiere_permiso_documentos_subir(self):
        self.client.force_login(self.dirigente_a)
        resp = self.client.get(reverse('core:sindiapp_documento_subir'))
        self.assertEqual(resp.status_code, 403)

    def test_subir_jpg_crea_documento_asociado_a_empresa(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._jpg_file()},
        )
        self.assertEqual(resp.status_code, 302)
        doc = DocumentoSindicato.objects.get(empresa=self.empresa_a)
        self.assertEqual(doc.empresa, self.empresa_a)
        self.assertEqual(doc.subido_por, self.tesoreria_a)
        self.assertEqual(doc.tipo_archivo, 'JPG')

    def test_subir_png_crea_documento(self):
        self.client.force_login(self.admin_a)
        self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._png_file()},
        )
        self.assertTrue(
            DocumentoSindicato.objects.filter(empresa=self.empresa_a, tipo_archivo='PNG').exists()
        )

    def test_subir_pdf_crea_documento(self):
        self.client.force_login(self.admin_a)
        self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._pdf_file()},
        )
        self.assertTrue(
            DocumentoSindicato.objects.filter(empresa=self.empresa_a, tipo_archivo='PDF').exists()
        )

    def test_subir_extension_invalida_rechaza(self):
        self.client.force_login(self.admin_a)
        archivo = SimpleUploadedFile('mal.txt', b'datos', content_type='text/plain')
        resp = self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': archivo},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(DocumentoSindicato.objects.filter(empresa=self.empresa_a).count(), 0)

    def test_subir_redirige_a_revisar(self):
        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._jpg_file()},
        )
        self.assertEqual(resp.status_code, 302)
        doc = DocumentoSindicato.objects.get(empresa=self.empresa_a)
        self.assertIn(str(doc.pk), resp.url)

    def test_subir_crea_auditoria(self):
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._jpg_file()},
        )
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='SUBIR_DOCUMENTO',
            ).exists()
        )

    # ---- Revisar ----

    def test_revisar_no_permite_ver_documento_otro_tenant(self):
        doc_b = DocumentoSindicato.objects.create(
            empresa=self.empresa_b,
            subido_por=self.admin_b,
            nombre_archivo='b.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
        )
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindiapp_documento_revisar', args=[doc_b.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_revisar_muestra_datos_extraidos(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='factura.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={'rut': '12.345.678-9', 'total': '100000', 'confianza': 'MEDIA'},
        )
        self.client.force_login(self.admin_a)
        resp = self.client.get(reverse('core:sindiapp_documento_revisar', args=[doc.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, '12.345.678-9')

    # ---- Confirmar ----

    def test_confirmar_crea_movimiento(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={'total': '50000', 'confianza': 'ALTA'},
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(
            reverse('core:sindiapp_documento_confirmar', args=[doc.pk]),
            data={
                'socio': self.socio_a.pk,
                'tipo_beneficio': self.benef_a.pk,
                'periodo': '2026-06',
                'monto': '50000',
                'observacion': 'Test confirmar',
            },
        )
        self.assertEqual(resp.status_code, 302)
        doc.refresh_from_db()
        self.assertEqual(doc.estado, DocumentoSindicato.ESTADO_CONFIRMADO)
        self.assertIsNotNone(doc.movimiento_creado)
        mov = doc.movimiento_creado
        self.assertEqual(mov.empresa, self.empresa_a)
        self.assertEqual(mov.fuente, MovimientoSindicato.FUENTE_DOCUMENTO)
        self.assertEqual(int(mov.monto), 50000)

    def test_confirmar_crea_auditoria(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_aud.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_confirmar', args=[doc.pk]),
            data={
                'socio': self.socio_a.pk,
                'tipo_beneficio': self.benef_a.pk,
                'periodo': '2026-06',
                'monto': '30000',
            },
        )
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='CONFIRMAR_DOCUMENTO',
            ).exists()
        )

    def test_confirmar_monto_invalido_no_crea_movimiento(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_inv.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_confirmar', args=[doc.pk]),
            data={
                'socio': self.socio_a.pk,
                'tipo_beneficio': self.benef_a.pk,
                'periodo': '2026-06',
                'monto': '0',
            },
        )
        doc.refresh_from_db()
        self.assertIsNone(doc.movimiento_creado)

    # ---- Rechazar ----

    def test_rechazar_no_crea_movimiento(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_rech.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.tesoreria_a)
        resp = self.client.post(
            reverse('core:sindiapp_documento_rechazar', args=[doc.pk]),
            data={'motivo': 'Documento ilegible'},
        )
        self.assertEqual(resp.status_code, 302)
        doc.refresh_from_db()
        self.assertEqual(doc.estado, DocumentoSindicato.ESTADO_RECHAZADO)
        self.assertIsNone(doc.movimiento_creado)
        self.assertEqual(MovimientoSindicato.objects.filter(empresa=self.empresa_a).count(), 0)

    def test_rechazar_sin_motivo_no_rechaza(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_rech_empty.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_rechazar', args=[doc.pk]),
            data={'motivo': ''},
        )
        doc.refresh_from_db()
        self.assertNotEqual(doc.estado, DocumentoSindicato.ESTADO_RECHAZADO)

    def test_rechazar_crea_auditoria(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_rech_aud.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_rechazar', args=[doc.pk]),
            data={'motivo': 'Motivo rechazo test'},
        )
        self.assertTrue(
            AuditoriaSindicato.objects.filter(
                empresa=self.empresa_a,
                accion='RECHAZAR_DOCUMENTO',
            ).exists()
        )

    def test_dirigente_no_puede_confirmar(self):
        doc = DocumentoSindicato.objects.create(
            empresa=self.empresa_a,
            subido_por=self.admin_a,
            nombre_archivo='doc_dir.jpg',
            tipo_archivo='JPG',
            estado=DocumentoSindicato.ESTADO_EN_REVISION,
            datos_extraidos={},
        )
        self.client.force_login(self.dirigente_a)
        resp = self.client.post(
            reverse('core:sindiapp_documento_confirmar', args=[doc.pk]),
            data={
                'socio': self.socio_a.pk,
                'tipo_beneficio': self.benef_a.pk,
                'periodo': '2026-06',
                'monto': '10000',
            },
        )
        self.assertEqual(resp.status_code, 403)
        doc.refresh_from_db()
        self.assertIsNone(doc.movimiento_creado)

    # ---- OCR service (sin OCR instalado → demo) ----

    def test_ocr_demo_extrae_datos_basicos(self):
        from core.services.sindicato_ocr import DemoOCRProvider
        provider = DemoOCRProvider()
        texto = provider.extract_text('fake_path.jpg')
        datos = provider.extract_data(texto)
        self.assertIn('12.345.678-9', datos.rut)
        self.assertTrue(int(datos.total) > 0)
        self.assertIn(datos.beneficio_sugerido, ['Gas', ''])

    def test_ocr_parse_rut_en_texto(self):
        from core.services.sindicato_ocr import _parse_texto
        texto = "Nombre: Juan Perez\nRUT: 12.345.678-9\nFactura N°1234\nTotal: $50.000"
        datos = _parse_texto(texto)
        self.assertEqual(datos.rut, '12.345.678-9')
        self.assertEqual(datos.numero_documento, '1234')
        self.assertEqual(datos.total, '50000')

    def test_subir_archivo_demasiado_grande_rechaza(self):
        self.client.force_login(self.admin_a)
        contenido_grande = b'X' * (11 * 1024 * 1024)
        archivo = SimpleUploadedFile('grande.jpg', contenido_grande, content_type='image/jpeg')
        resp = self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': archivo},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(DocumentoSindicato.objects.filter(empresa=self.empresa_a).count(), 0)

    def test_auditoria_subir_guarda_confianza_no_proveedor(self):
        self.client.force_login(self.tesoreria_a)
        self.client.post(
            reverse('core:sindiapp_documento_subir'),
            data={'archivo': self._jpg_file('audit_conf.jpg')},
        )
        audit = AuditoriaSindicato.objects.filter(
            empresa=self.empresa_a, accion='SUBIR_DOCUMENTO'
        ).last()
        self.assertIsNotNone(audit)
        self.assertIn('confianza=', audit.resumen)
        self.assertNotIn('proveedor=', audit.resumen)

    def test_ocr_error_marca_documento_como_error(self):
        """Si el proveedor OCR falla, el documento queda en estado ERROR."""
        from unittest.mock import patch

        self.client.force_login(self.tesoreria_a)
        with patch(
            'core.services.sindicato_ocr.procesar_documento',
            side_effect=RuntimeError('fallo OCR simulado'),
        ):
            resp = self.client.post(
                reverse('core:sindiapp_documento_subir'),
                data={'archivo': self._jpg_file('error_ocr.jpg')},
            )
        self.assertEqual(resp.status_code, 302)
        doc = DocumentoSindicato.objects.get(empresa=self.empresa_a, nombre_archivo='error_ocr.jpg')
        self.assertEqual(doc.estado, DocumentoSindicato.ESTADO_ERROR)
        self.assertIn('fallo OCR simulado', doc.error_mensaje)
