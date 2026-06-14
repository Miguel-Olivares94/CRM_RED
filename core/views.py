from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LogoutView as DjangoLogoutView
from django.core.exceptions import PermissionDenied
from django.db.models import Q, Count, Sum
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import CreateView, ListView, TemplateView, DetailView, UpdateView, DeleteView
from django.http import Http404, HttpResponse, HttpResponseRedirect, JsonResponse, HttpResponseForbidden
from django.utils.translation import activate
from datetime import datetime, timedelta
import json
import io
import csv
import hashlib
from decimal import Decimal, InvalidOperation
import re

from .forms import (
    ClienteForm, ContactoForm, OportunidadForm, LlamadaForm,
    MetaVentasForm, ComisionForm, SeguimientoForm, EmailAuthenticationForm,
    CreateEjecutivoForm, SocioSindicatoForm, TipoBeneficioSindicatoForm,
    MovimientoSindicatoForm,
)
from .models import (
    Cliente, Contacto, Oportunidad, Llamada,
    MetaVentas, Comision, Seguimiento, UserProfile, CampoPersonalizado,
    Empresa, SocioSindicato, TipoBeneficioSindicato, MovimientoSindicato,
    ConsolidadoMensualSindicato, ConsolidadoDetalleSindicato,
    AuditoriaSindicato,
)
from .mixins import (
    AdminOnlyMixin, EjecutivoOrAdminMixin, ManagerOnlyMixin,
    ClienteQuerysetFilterMixin, OportunidadQuerysetFilterMixin,
    ContactoQuerysetFilterMixin, LlamadaQuerysetFilterMixin,
    SeguimientoQuerysetFilterMixin, ComisionQuerysetFilterMixin,
    TenantFilterMixin, get_user_role
)
from .services.sindicato_consolidado import (
    ConsolidadoBloqueadoError,
    ConsolidadoNoExisteError,
    cerrar_consolidado_periodo,
    generar_o_recalcular_consolidado,
    recalcular_consolidado_abierto,
)
from .services.sindicato_exportacion import ConsolidadoExportacionError, exportar_consolidado_excel


def _build_supervisor_ejecutivos_json():
    """Devuelve JSON {supervisor_nombre: [ejecutivo_nombre, ...]} para el cascade del formulario."""
    from django.contrib.auth import get_user_model as _get_user_model
    _User = _get_user_model()

    def _nombre(u):
        return u.get_full_name().strip() or u.username

    mapping = {}
    todos_ejecutivos = []

    supervisores = _User.objects.filter(
        groups__name__in=['Manager', 'Admin']
    ).prefetch_related('subordinados__user').order_by('last_name', 'first_name')

    for sup in supervisores:
        n_sup = _nombre(sup)
        ejecs = sorted(
            [_nombre(p.user) for p in sup.subordinados.select_related('user').all()],
            key=str.lower
        )
        mapping[n_sup] = ejecs
        for e in ejecs:
            if e not in todos_ejecutivos:
                todos_ejecutivos.append(e)

    return json.dumps(mapping), json.dumps(sorted(todos_ejecutivos, key=str.lower))


# ==================== HOME / INDEX ====================
class IndexView(LoginRequiredMixin, View):
    """
    Vista de inicio que redirija según el rol del usuario:
    - Admin/Superadmin → Dashboard completo
    - Vendor/Ejecutivo → Dashboard de vendedor
    - Sin autenticar → Login
    """
    login_url = 'core:login'
    
    def get(self, request, *args, **kwargs):
        user = request.user

        if user.is_superuser or user.groups.filter(name='Admin').exists():
            return redirect('core:dashboard')

        if user.groups.filter(name='Manager').exists():
            return redirect('core:dashboard')

        if user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists():
            return redirect('core:ejecutivo_clientes')

        return redirect('core:login')


# ==================== DASHBOARD ====================
class DashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard para administradores y managers"""
    template_name = "core/dashboard.html"
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar que sea Admin o Manager"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        # Si no es admin ni manager, redirigir
        if 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return redirect('core:ejecutivo_clientes')
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Verificar si es Manager (tiene subordinados)
        user = self.request.user
        es_manager = False
        subordinados_y_yo = []
        
        # Intentar obtener el UserProfile
        try:
            from .models import UserProfile
            profile = user.profile
            es_manager = profile.role == 'MANAGER'
            if es_manager:
                # Obtener IDs de los subordinados + el manager mismo
                subordinados_y_yo = list(user.subordinados.values_list('user_id', flat=True))
                subordinados_y_yo.append(user.id)  # Incluir al manager
        except:
            pass
        
        # Decidir qué datos filtrar
        if es_manager and subordinados_y_yo:
            # MANAGER ve a sus subordinados + a sí mismo
            oportunidades_query = Oportunidad.objects.filter(usuario_id__in=subordinados_y_yo)
            clientes_query = Cliente.objects.filter(usuario_asignado_id__in=subordinados_y_yo)
            llamadas_query = Llamada.objects.filter(oportunidad__usuario_id__in=subordinados_y_yo)
        else:
            # ADMIN ve todos
            oportunidades_query = Oportunidad.objects.all()
            clientes_query = Cliente.objects.all()
            llamadas_query = Llamada.objects.all()
        
        # KPIs Generales
        context["total_clientes"] = clientes_query.count()
        context["total_oportunidades"] = oportunidades_query.filter(estado='ABIERTA').count()
        context["total_llamadas"] = llamadas_query.count()
        context["es_manager"] = es_manager
        
        # Oportunidades por etapa
        context["oportunidades_por_etapa"] = oportunidades_query.values('etapa').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Total líneas
        context["total_lineas"] = clientes_query.aggregate(
            total=Sum('q_total_lineas')
        )['total'] or 0
        
        # Últimas oportunidades
        context["ultimas_oportunidades"] = oportunidades_query.select_related(
            'cliente', 'usuario'
        ).order_by('-created_at')[:10]
        
        # Seguimientos pendientes
        context["seguimientos_pendientes"] = Seguimiento.objects.filter(
            estado='PENDIENTE',
            oportunidad__in=oportunidades_query
        ).select_related('oportunidad', 'asignado_a').order_by('fecha_vencimiento')[:10]
        
        # ==================== DATOS PARA GRÁFICOS ====================

        # 1. TOP 10 CLIENTES POR TOTAL LÍNEAS
        top_clientes = list(
            clientes_query.filter(q_total_lineas__gt=0)
            .order_by('-q_total_lineas')
            .values('nombre_empresa', 'q_total_lineas')[:10]
        )
        context['lineas_clientes_labels'] = json.dumps([c['nombre_empresa'] for c in top_clientes])
        context['lineas_clientes_data']   = json.dumps([c['q_total_lineas'] or 0 for c in top_clientes])

        # 2. DISTRIBUCIÓN DE TIPOS DE LÍNEAS (BAM / M2M / VOZ-GPS)
        totales = clientes_query.aggregate(
            total_bam=Sum('bam'),
            total_m2m=Sum('m2m'),
            total_voz=Sum('voz'),
        )
        context['tipos_lineas_labels'] = json.dumps(['BAM', 'M2M', 'VOZ / GPS'])
        context['tipos_lineas_data']   = json.dumps([
            totales['total_bam'] or 0,
            totales['total_m2m'] or 0,
            totales['total_voz'] or 0,
        ])

        return context


# ==================== SUBORDINADOS MANAGEMENT (para Managers y Admins) ====================
class SubordinadosManagementView(LoginRequiredMixin, TemplateView):
    """Dashboard de gestión de vendedores para managers y administradores"""
    template_name = "core/subordinados_management.html"
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Verificar que sea Admin o Manager"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Verificar si es manager o admin
        user_groups = user.groups.values_list('name', flat=True)
        is_admin = user.is_superuser or 'Admin' in user_groups
        is_manager = 'Manager' in user_groups
        
        # Para managers: mostrar sus subordinados. Para admins: mostrar todos los usuarios con rol Vendedor
        if is_manager:
            subordinados_profiles = user.subordinados.all()
        elif is_admin:
            # Obtener todos los usuarios en grupo Vendedor
            from django.contrib.auth import get_user_model
            from django.contrib.auth.models import Group
            User = get_user_model()
            vendedor_group = Group.objects.filter(name='Vendedor').first()
            subordinados_profiles = []
            if vendedor_group:
                for user_obj in vendedor_group.user_set.all():
                    if hasattr(user_obj, 'profile'):
                        subordinados_profiles.append(user_obj.profile)
        else:
            subordinados_profiles = []
        
        subordinados_data = []
        
        for profile in subordinados_profiles:
            subordinado_user = profile.user
            
            # Clientes asignados
            clientes = Cliente.objects.filter(usuario_asignado=subordinado_user)
            total_clientes = clientes.count()
            
            # Oportunidades
            oportunidades = Oportunidad.objects.filter(usuario=subordinado_user)
            total_oportunidades_abiertas = oportunidades.filter(estado='ABIERTA').count()
            total_oportunidades_ganadas = oportunidades.filter(estado='GANADA').count()
            total_oportunidades_perdidas = oportunidades.filter(estado='PERDIDA').count()
            
            # Cartera total
            cartera_total = oportunidades.filter(estado='ABIERTA').aggregate(
                total=Sum('monto')
            )['total'] or 0
            
            # Total líneas del vendedor
            total_lineas = clientes.aggregate(
                total=Sum('q_total_lineas')
            )['total'] or 0
            
            # Oportunidades por etapa
            oportunidades_por_etapa = oportunidades.values('etapa').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Llamadas realizadas
            llamadas_count = Llamada.objects.filter(
                oportunidad__usuario=subordinado_user
            ).count()
            
            # Últimas oportunidades
            ultimas_oportunidades = oportunidades.select_related('cliente').order_by('-created_at')[:5]
            
            # Seguimientos pendientes
            seguimientos_pendientes = Seguimiento.objects.filter(
                estado='PENDIENTE',
                oportunidad__usuario=subordinado_user
            ).select_related('oportunidad').order_by('fecha_vencimiento')[:5]
            
            subordinados_data.append({
                'profile': profile,
                'user': subordinado_user,
                'total_clientes': total_clientes,
                'total_oportunidades_abiertas': total_oportunidades_abiertas,
                'total_oportunidades_ganadas': total_oportunidades_ganadas,
                'total_oportunidades_perdidas': total_oportunidades_perdidas,
                'cartera_total': cartera_total,
                'total_lineas': total_lineas,
                'oportunidades_por_etapa': oportunidades_por_etapa,
                'llamadas_count': llamadas_count,
                'ultimas_oportunidades': ultimas_oportunidades,
                'seguimientos_pendientes': seguimientos_pendientes,
            })
        
        context['subordinados_data'] = subordinados_data
        context['total_subordinados'] = len(subordinados_data)
        
        # Estadísticas consolidadas del equipo
        if is_manager:
            team_users_ids = [profile.user_id for profile in subordinados_profiles]
            team_users_ids.append(user.id)  # Incluir al manager
        elif is_admin:
            # Para admins, mostrar todos los usuarios
            team_users_ids = list(User.objects.values_list('id', flat=True))
        else:
            team_users_ids = []
        
        context['team_total_clientes'] = Cliente.objects.filter(
            usuario_asignado_id__in=team_users_ids
        ).count() if team_users_ids else 0
        
        context['team_total_oportunidades'] = Oportunidad.objects.filter(
            usuario_id__in=team_users_ids,
            estado='ABIERTA'
        ).count() if team_users_ids else 0
        
        context['team_cartera_total'] = Oportunidad.objects.filter(
            usuario_id__in=team_users_ids,
            estado='ABIERTA'
        ).aggregate(total=Sum('monto'))['total'] or 0 if team_users_ids else 0
        
        context['team_oportunidades_ganadas'] = Oportunidad.objects.filter(
            usuario_id__in=team_users_ids,
            estado='GANADA'
        ).count() if team_users_ids else 0
        
        context['team_total_lineas'] = Cliente.objects.filter(
            usuario_asignado_id__in=team_users_ids
        ).aggregate(total=Sum('q_total_lineas'))['total'] or 0 if team_users_ids else 0
        
        return context


# ==================== CREAR EJECUTIVO (para Managers) ====================
class CreateEjecutivoView(ManagerOnlyMixin, View):
    """Vista para que managers creen nuevos ejecutivos"""
    template_name = "core/crear_ejecutivo.html"
    
    def get(self, request):
        form = CreateEjecutivoForm()
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        form = CreateEjecutivoForm(request.POST)
        
        if form.is_valid():
            try:
                from django.contrib.auth import get_user_model
                from .models import UserProfile
                
                User = get_user_model()
                
                # Generar contraseña segura
                import secrets
                import string
                password = ''.join(
                    secrets.choice(string.ascii_letters + string.digits + '!@#$%^&*')
                    for _ in range(16)
                )
                
                # Crear usuario
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    password=password
                )

                # Agregar al grupo Django 'Ejecutivo' para que tenga acceso correcto
                from django.contrib.auth.models import Group
                grupo_ejecutivo, _ = Group.objects.get_or_create(name='Ejecutivo')
                user.groups.add(grupo_ejecutivo)

                # Crear UserProfile como EJECUTIVO bajo este manager
                UserProfile.objects.create(
                    user=user,
                    role='EJECUTIVO',
                    supervisor=request.user
                )
                
                # Respuesta de éxito
                return render(request, self.template_name, {
                    'form': CreateEjecutivoForm(),
                    'success': True,
                    'nuevo_usuario': {
                        'nombre': user.get_full_name(),
                        'email': user.email,
                        'username': user.username,
                        'password': password
                    }
                })
            
            except Exception as e:
                form.add_error(None, f"Error al crear ejecutivo: {str(e)}")
        
        return render(request, self.template_name, {'form': form})


# ==================== ASIGNAR EQUIPO ====================
class AsignarEquipoView(LoginRequiredMixin, View):
    """Admin y Manager asignan ejecutivos a supervisores."""
    template_name = "core/asignar_equipo.html"

    def _check_permission(self, request):
        u = request.user
        return u.is_superuser or u.groups.filter(name__in=['Admin', 'Manager']).exists()

    def dispatch(self, request, *args, **kwargs):
        if not self._check_permission(request):
            return HttpResponseForbidden("Sin permiso.")
        return super().dispatch(request, *args, **kwargs)

    def _context(self, request):
        from django.contrib.auth import get_user_model as _gum
        from django.contrib.auth.models import Group
        _User = _gum()
        u = request.user
        es_admin = u.is_superuser or u.groups.filter(name__in=['Admin']).exists()

        managers = _User.objects.filter(
            groups__name__in=['Manager', 'Admin']
        ).order_by('last_name', 'first_name')

        ejecutivos_qs = _User.objects.filter(
            groups__name__in=['Ejecutivo', 'Vendedor']
        ).select_related('profile').order_by('last_name', 'first_name').distinct()

        # Manager solo ve a sus propios subordinados
        if not es_admin:
            ejecutivos_qs = ejecutivos_qs.filter(profile__supervisor=u)

        ejecutivos = []
        for ej in ejecutivos_qs:
            sup = getattr(getattr(ej, 'profile', None), 'supervisor', None)
            ejecutivos.append({'user': ej, 'supervisor': sup})

        return {
            'ejecutivos': ejecutivos,
            'managers': managers,
            'es_admin': es_admin,
        }

    def get(self, request):
        from django.contrib import messages
        ctx = self._context(request)
        return render(request, self.template_name, ctx)

    def post(self, request):
        from django.contrib import messages
        from django.contrib.auth import get_user_model as _gum
        _User = _gum()
        u = request.user
        es_admin = u.is_superuser or u.groups.filter(name__in=['Admin']).exists()

        ejecutivo_id = request.POST.get('ejecutivo_id')
        supervisor_id = request.POST.get('supervisor_id') or None

        try:
            profile = UserProfile.objects.get(user_id=ejecutivo_id)
            # Manager solo puede operar sobre sus subordinados
            if not es_admin and profile.supervisor != u:
                return HttpResponseForbidden("Sin permiso sobre este ejecutivo.")
            profile.supervisor_id = supervisor_id if supervisor_id else None
            profile.save()
            messages.success(request, "Asignación actualizada correctamente.")
        except UserProfile.DoesNotExist:
            messages.error(request, "Ejecutivo no encontrado.")

        return redirect('core:asignar_equipo')


# ==================== CLIENTE ====================
class ClienteListView(LoginRequiredMixin, ClienteQuerysetFilterMixin, ListView):
    model = Cliente
    template_name = 'core/cliente_list.html'
    context_object_name = "clientes"
    paginate_by = 20
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        # Primero aplica el filtro del mixin (admin ve todos, ejecutivo ve solo sus)
        queryset = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        if q:
            queryset = queryset.filter(
                Q(nombre_empresa__icontains=q) | 
                Q(rut__icontains=q) |
                Q(sector__icontains=q)
            )
        estado = self.request.GET.get("estado", "").strip()
        if estado:
            queryset = queryset.filter(estado=estado)
        con_lineas = self.request.GET.get("con_lineas", "")
        if con_lineas == "1":
            queryset = queryset.filter(q_total_lineas__gt=0)
        user_groups = self.request.user.groups.values_list('name', flat=True)
        is_admin = self.request.user.is_superuser or 'Admin' in user_groups
        # Filtro por supervisor (solo admin): muestra equipo completo del supervisor
        manager_id = self.request.GET.get("supervisor", "").strip()
        if manager_id and is_admin:
            subordinados_ids = list(
                UserProfile.objects.filter(supervisor_id=manager_id).values_list('user_id', flat=True)
            )
            subordinados_ids.append(int(manager_id))
            queryset = queryset.filter(usuario_asignado_id__in=subordinados_ids)
        # Filtro por ejecutivo individual
        ejecutivo_id = self.request.GET.get("ejecutivo", "").strip()
        if ejecutivo_id:
            queryset = queryset.filter(usuario_asignado_id=ejecutivo_id)

        # Filtro por segmento
        segmento = self.request.GET.get("segmento", "").strip()
        if segmento:
            queryset = queryset.filter(segmento__iexact=segmento)

        # Filtro mínimo de líneas
        min_lineas = self.request.GET.get("min_lineas", "").strip()
        if min_lineas.isdigit():
            queryset = queryset.filter(q_total_lineas__gte=int(min_lineas))

        # Ordenamiento
        orden = self.request.GET.get("orden", "")
        orden_map = {
            "lineas_desc": "-q_total_lineas",
            "lineas_asc":  "q_total_lineas",
            "nombre":      "nombre_empresa",
            "reciente":    "-created_at",
            "antiguo":     "created_at",
        }
        return queryset.order_by(orden_map.get(orden, "-q_total_lineas"))
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_role'] = get_user_role(self.request.user)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user_groups = self.request.user.groups.values_list('name', flat=True)
        is_admin = self.request.user.is_superuser or 'Admin' in user_groups
        is_manager = 'Manager' in user_groups
        if is_admin:
            # Admin ve dos filtros: supervisores y ejecutivos/vendedores
            context['supervisores'] = User.objects.filter(
                groups__name='Manager'
            ).distinct().order_by('first_name', 'last_name')
            context['ejecutivos'] = User.objects.filter(
                groups__name__in=['Ejecutivo', 'Vendedor']
            ).distinct().order_by('first_name', 'last_name')
        elif is_manager:
            # Manager ve sus vendedores/ejecutivos subordinados
            subordinados = self.request.user.subordinados.select_related('user').order_by(
                'user__first_name', 'user__last_name'
            )
            context['ejecutivos'] = [p.user for p in subordinados]

        # Segmentos disponibles para el filtro
        context['segmentos'] = (
            Cliente.objects.filter(segmento__isnull=False)
            .exclude(segmento="")
            .values_list('segmento', flat=True)
            .distinct()
            .order_by('segmento')
        )
        return context


class ClienteCreateView(EjecutivoOrAdminMixin, CreateView):
    """Admins pueden crear clientes y asignarlo a cualquiera. Ejecutivos crean clientes auto-asignados a ellos."""
    model = Cliente
    form_class = ClienteForm
    template_name = "core/cliente_form.html"
    success_url = reverse_lazy("core:cliente_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        user = self.request.user
        nombre = user.get_full_name().strip() or user.username
        try:
            if user.profile.es_manager:
                initial['supervisor'] = nombre
        except Exception:
            if user.groups.filter(name='Manager').exists():
                initial['supervisor'] = nombre
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Cliente"
        context["accion"] = "crear"
        context["supervisor_ejecutivos_json"], context["todos_ejecutivos_json"] = _build_supervisor_ejecutivos_json()
        try:
            context["es_manager"]   = self.request.user.profile.es_manager
            context["es_ejecutivo"] = self.request.user.profile.es_ejecutivo
        except Exception:
            context["es_manager"]   = self.request.user.groups.filter(name='Manager').exists()
            context["es_ejecutivo"] = self.request.user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists()
        return context

    def form_valid(self, form):
        """Asigna el cliente automáticamente según el rol del usuario"""
        user = self.request.user

        # Si es ejecutivo, auto-asigna a él mismo
        if user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists():
            form.instance.usuario_asignado = user

        # Marcar como prospectado manualmente y registrar quién lo creó
        form.instance.origen = 'MANUAL'
        form.instance.creado_por = user

        # Auto-asignar empresa del usuario (Capa 1 de defensa anti-fuga)
        try:
            form.instance.empresa = user.profile.empresa
        except Exception:
            pass

        # Guardar valores de campos personalizados en datos_extra
        datos_extra = form.instance.datos_extra or {}
        for key, value in form.cleaned_data.items():
            if key.startswith('extra__'):
                clave = key[len('extra__'):]
                # Serializar fecha a string para JSON
                if hasattr(value, 'isoformat'):
                    value = value.isoformat()
                datos_extra[clave] = value
        form.instance.datos_extra = datos_extra

        return super().form_valid(form)


class ClienteDetailView(EjecutivoOrAdminMixin, ClienteQuerysetFilterMixin, DetailView):
    model = Cliente
    template_name = "core/cliente_detail.html"
    context_object_name = "cliente"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente = self.get_object()
        context["contactos"] = cliente.contactos.all()
        context["oportunidades"] = cliente.oportunidades.filter(estado='ABIERTA')
        context["valor_cartera"] = cliente.valor_cartera
        context['user_role'] = get_user_role(self.request.user)
        # Pasar campos personalizados de la empresa para mostrar datos_extra
        try:
            empresa = cliente.empresa
            if empresa:
                campos_qs = CampoPersonalizado.objects.filter(
                    empresa=empresa,
                    entidad=CampoPersonalizado.ENTIDAD_CLIENTE,
                    activo=True,
                ).order_by('orden', 'nombre')
                datos_extra = cliente.datos_extra or {}
                # Prepara lista [{nombre, tipo, valor}] solo para campos con valor
                context['campos_extra_valores'] = [
                    {
                        'nombre': c.nombre,
                        'tipo': c.tipo,
                        'valor': datos_extra.get(c.clave, ''),
                    }
                    for c in campos_qs
                    if datos_extra.get(c.clave) not in (None, '', {}, [])
                ]
            else:
                context['campos_extra_valores'] = []
        except Exception:
            context['campos_extra_valores'] = []
        return context
    
    def get_object(self, queryset=None):
        """Verifica que el usuario tenga permiso de ver este cliente"""
        obj = super().get_object(queryset)
        user = self.request.user
        # Si es ejecutivo, verifica que sea su cliente asignado
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para ver este cliente")
        return obj


class ClienteUpdateView(EjecutivoOrAdminMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "core/cliente_form.html"
    success_url = reverse_lazy("core:cliente_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        # Pasar empresa del usuario para inyectar campos personalizados
        try:
            kwargs['empresa'] = self.request.user.profile.empresa
        except Exception:
            kwargs['empresa'] = None
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Cliente"
        context["accion"] = "editar"
        context['user_role'] = get_user_role(self.request.user)
        context["supervisor_ejecutivos_json"], context["todos_ejecutivos_json"] = _build_supervisor_ejecutivos_json()
        try:
            context["es_manager"]   = self.request.user.profile.es_manager
            context["es_ejecutivo"] = self.request.user.profile.es_ejecutivo
        except Exception:
            context["es_manager"]   = self.request.user.groups.filter(name='Manager').exists()
            context["es_ejecutivo"] = self.request.user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists()
        return context

    def get_object(self, queryset=None):
        """Verifica que el usuario tenga permiso de editar este cliente"""
        obj = super().get_object(queryset)
        user = self.request.user
        # Si es ejecutivo, verifica que sea su cliente asignado
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para editar este cliente")
        return obj

    def form_valid(self, form):
        """Guarda valores de campos personalizados en datos_extra."""
        datos_extra = form.instance.datos_extra or {}
        for key, value in form.cleaned_data.items():
            if key.startswith('extra__'):
                clave = key[len('extra__'):]
                if hasattr(value, 'isoformat'):
                    value = value.isoformat()
                datos_extra[clave] = value
        form.instance.datos_extra = datos_extra
        return super().form_valid(form)


class ClienteDeleteView(AdminOnlyMixin, DeleteView):
    """Solo admins pueden eliminar clientes"""
    model = Cliente
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("core:cliente_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["objeto_nombre"] = self.object.nombre_empresa
        context["tipo_objeto"] = "Cliente"
        return context


# ==================== CONTACTO ====================
class ContactoListView(LoginRequiredMixin, ContactoQuerysetFilterMixin, ListView):
    model = Contacto
    template_name = "core/contacto_list.html"
    context_object_name = "contactos"
    paginate_by = 20
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get("q", "").strip()
        cliente_id = self.request.GET.get("cliente", "").strip()
        
        if q:
            queryset = queryset.filter(
                Q(nombre__icontains=q) |
                Q(email__icontains=q) |
                Q(cliente__nombre_empresa__icontains=q)
            )
        if cliente_id:
            queryset = queryset.filter(cliente_id=cliente_id)
            
        return queryset.select_related('cliente').order_by('cliente', 'orden')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Para cada contacto, agregar su oportunidad asociada
        for contacto in context['contactos']:
            # Obtener la primera oportunidad abierta del cliente
            oportunidad = Oportunidad.objects.filter(
                cliente=contacto.cliente,
                estado='ABIERTA'
            ).first()
            contacto.oportunidad_abierta = oportunidad
        
        context['user_role'] = get_user_role(self.request.user)
        return context


class ContactoDetailView(EjecutivoOrAdminMixin, ContactoQuerysetFilterMixin, DetailView):
    model = Contacto
    template_name = "core/contacto_detail.html"
    context_object_name = "contacto"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contacto = self.get_object()
        context['llamadas'] = contacto.llamadas.select_related('oportunidad').order_by('-fecha_hora')[:10]
        context['oportunidades'] = Oportunidad.objects.filter(cliente=contacto.cliente, estado='ABIERTA')
        context['user_role'] = get_user_role(self.request.user)
        return context


class ContactoUpdateView(EjecutivoOrAdminMixin, UpdateView):
    model = Contacto
    form_class = ContactoForm
    template_name = "core/contacto_form.html"
    success_url = reverse_lazy("core:contacto_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Contacto"
        context["accion"] = "editar"
        return context

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para editar este contacto")
        return obj


class ContactoDeleteView(AdminOnlyMixin, DeleteView):
    model = Contacto
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("core:contacto_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["objeto_nombre"] = self.object.nombre
        context["tipo_objeto"] = "Contacto"
        return context


class ContactoCreateView(EjecutivoOrAdminMixin, CreateView):
    model = Contacto
    form_class = ContactoForm
    template_name = "core/contacto_form.html"
    success_url = reverse_lazy("core:contacto_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        cliente_id = self.request.GET.get('cliente')
        if cliente_id:
            initial['cliente'] = cliente_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Contacto"
        context["accion"] = "crear"
        return context

    def form_valid(self, form):
        user = self.request.user
        cliente = form.cleaned_data['cliente']

        if user.groups.filter(name='Ejecutivo').exists():
            if cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para agregar contactos a este cliente")

        return super().form_valid(form)


# ==================== OPORTUNIDAD ====================
class OportunidadListView(LoginRequiredMixin, OportunidadQuerysetFilterMixin, ListView):
    model = Oportunidad
    template_name = "core/oportunidad_list.html"
    context_object_name = "oportunidades"
    paginate_by = 20
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        # Primero aplica el filtro del mixin (admin ve todas, ejecutivo solo sus)
        queryset = super().get_queryset()
        
        # Filtros adicionales
        estado = self.request.GET.get("estado", "").strip()
        etapa = self.request.GET.get("etapa", "").strip()
        usuario_id = self.request.GET.get("usuario", "").strip()
        q = self.request.GET.get("q", "").strip()
        
        if estado:
            queryset = queryset.filter(estado=estado)
        if etapa:
            queryset = queryset.filter(etapa=etapa)
        # Los ejecutivos no pueden filtrar por usuario (solo ven las suyas)
        if usuario_id and (self.request.user.is_superuser or self.request.user.groups.filter(name='Admin').exists()):
            queryset = queryset.filter(usuario_id=usuario_id)
        if q:
            queryset = queryset.filter(
                Q(cliente__nombre_empresa__icontains=q) |
                Q(cliente__rut__icontains=q) |
                Q(productos__icontains=q)
            )
            
        return queryset.select_related('cliente', 'usuario').order_by('-fecha_creacion')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_role'] = get_user_role(self.request.user)
        return context


class OportunidadCreateView(EjecutivoOrAdminMixin, CreateView):
    model = Oportunidad
    form_class = OportunidadForm
    template_name = "core/oportunidad_form.html"
    success_url = reverse_lazy("core:oportunidad_list")

    def get_initial(self):
        initial = super().get_initial()
        cliente_id = self.request.GET.get('cliente')
        if cliente_id:
            initial['cliente'] = cliente_id
        return initial

    def dispatch(self, request, *args, **kwargs):
        # Ejecutivos sin clientes asignados no pueden crear oportunidades
        user = request.user
        es_ejecutivo = (
            not user.is_superuser and
            not user.groups.filter(name__in=['Admin', 'Manager']).exists()
        )
        if es_ejecutivo:
            tiene_clientes = Cliente.objects.filter(usuario_asignado=user).exists()
            if not tiene_clientes:
                from django.contrib import messages
                messages.warning(
                    request,
                    'No tienes clientes asignados. Contacta a tu manager para que te asigne clientes antes de crear oportunidades.'
                )
                return redirect('core:oportunidad_list')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nueva Oportunidad"
        context["accion"] = "crear"
        context['user_role'] = get_user_role(self.request.user)
        return context


class OportunidadDeleteView(AdminOnlyMixin, DeleteView):
    model = Oportunidad
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("core:oportunidad_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["objeto_nombre"] = f"{self.object.cliente.nombre_empresa} — {self.object.get_etapa_display()}"
        context["tipo_objeto"] = "Oportunidad"
        return context


class OportunidadDetailView(EjecutivoOrAdminMixin, OportunidadQuerysetFilterMixin, DetailView):
    model = Oportunidad
    template_name = "core/oportunidad_detail.html"
    context_object_name = "oportunidad"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        oportunidad = self.get_object()
        context["llamadas"] = oportunidad.llamadas.all().order_by('-fecha_hora')
        context["seguimientos"] = oportunidad.seguimientos.all().order_by('-fecha_creacion')
        context['user_role'] = get_user_role(self.request.user)
        return context
    
    def get_object(self, queryset=None):
        """Verifica que el usuario tenga permiso de ver esta oportunidad"""
        obj = super().get_object(queryset)
        user = self.request.user
        # Si es ejecutivo, verifica que el cliente sea asignado a él
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para ver esta oportunidad")
        return obj


class OportunidadUpdateView(EjecutivoOrAdminMixin, UpdateView):
    model = Oportunidad
    form_class = OportunidadForm
    template_name = "core/oportunidad_form.html"

    def get_form_kwargs(self):
        """Configura el formulario según el rol"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy("core:oportunidad_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Oportunidad"
        context["accion"] = "editar"
        context['user_role'] = get_user_role(self.request.user)
        return context

    def get_object(self, queryset=None):
        """Verifica que el usuario tenga permiso de editar esta oportunidad"""
        obj = super().get_object(queryset)
        user = self.request.user
        # Si es ejecutivo, verifica que el cliente sea asignado a él
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para editar esta oportunidad")
        return obj


# ==================== LLAMADA ====================
class LlamadaListView(LoginRequiredMixin, LlamadaQuerysetFilterMixin, ListView):
    model = Llamada
    template_name = "core/llamada_list.html"
    context_object_name = "llamadas"
    paginate_by = 20
    login_url = 'core:login'
    
    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        queryset = super().get_queryset()
        tipo = self.request.GET.get("tipo", "").strip()
        resultado = self.request.GET.get("resultado", "").strip()
        
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if resultado:
            queryset = queryset.filter(resultado=resultado)
            
        return queryset.select_related('oportunidad', 'contacto').order_by('-fecha_hora')


class LlamadaCreateView(LoginRequiredMixin, CreateView):
    model = Llamada
    form_class = LlamadaForm
    template_name = "core/llamada_form.html"
    success_url = reverse_lazy("core:llamada_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['contacto_id'] = self.request.GET.get('contacto')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Registrar Llamada/Contacto"
        context["accion"] = "crear"
        contacto_id = self.request.GET.get('contacto')
        if contacto_id:
            try:
                contacto_obj = Contacto.objects.select_related('cliente').get(pk=contacto_id)
                context['contacto_preseleccionado'] = contacto_obj
            except (Contacto.DoesNotExist, ValueError):
                pass
        from .chile_data import PRODUCTOS_CLARO_CHOICES
        context['productos_choices'] = PRODUCTOS_CLARO_CHOICES
        context['etapa_choices'] = Oportunidad.ETAPA_CHOICES
        return context

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        # Si el ejecutivo marcó "Abrir nueva oportunidad"
        if self.request.POST.get('abrir_oportunidad') == '1':
            etapa = self.request.POST.get('op_etapa', '').strip()
            if etapa:
                # Determinar el cliente desde el contacto
                cliente = None
                contacto_id = self.request.GET.get('contacto')
                if contacto_id:
                    try:
                        contacto_obj = Contacto.objects.select_related('cliente').get(pk=contacto_id)
                        cliente = contacto_obj.cliente
                    except (Contacto.DoesNotExist, ValueError):
                        pass
                if not cliente and form.cleaned_data.get('contacto'):
                    cliente = form.cleaned_data['contacto'].cliente
                if cliente:
                    from decimal import Decimal, InvalidOperation
                    from datetime import datetime
                    monto_raw = self.request.POST.get('op_monto', '').strip()
                    try:
                        monto = Decimal(monto_raw) if monto_raw else Decimal('0')
                    except InvalidOperation:
                        monto = Decimal('0')
                    prob_raw = self.request.POST.get('op_probabilidad', '').strip()
                    probabilidad = int(prob_raw) if prob_raw and prob_raw.isdigit() else 0
                    lineas_raw = self.request.POST.get('op_lineas', '').strip()
                    lineas = int(lineas_raw) if lineas_raw and lineas_raw.isdigit() else 0
                    fecha_raw = self.request.POST.get('op_fecha_cierre_estimada', '').strip()
                    try:
                        fecha_cierre = datetime.strptime(fecha_raw, '%Y-%m-%d').date() if fecha_raw else None
                    except ValueError:
                        fecha_cierre = None
                    oportunidad = Oportunidad.objects.create(
                        cliente=cliente,
                        usuario=self.request.user,
                        etapa=etapa,
                        estado='ABIERTA',
                        monto=monto,
                        moneda=self.request.POST.get('op_moneda', 'CLP'),
                        probabilidad=probabilidad,
                        productos=self.request.POST.get('op_productos', '') or None,
                        lineas=lineas,
                        fecha_cierre_estimada=fecha_cierre,
                        observaciones=self.request.POST.get('op_observaciones', '') or None,
                        creado_por=self.request.user,
                    )
                    form.instance.oportunidad = oportunidad
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("core:llamada_detail", kwargs={"pk": self.object.pk})


class LlamadaDetailView(LoginRequiredMixin, LlamadaQuerysetFilterMixin, DetailView):
    model = Llamada
    template_name = "core/llamada_detail.html"
    context_object_name = "llamada"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_role'] = get_user_role(self.request.user)
        return context


class LlamadaUpdateView(LoginRequiredMixin, UpdateView):
    model = Llamada
    form_class = LlamadaForm
    template_name = "core/llamada_form.html"
    success_url = reverse_lazy("core:llamada_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Gestión"
        context["accion"] = "editar"
        return context

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        user = self.request.user
        if user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists():
            if obj.oportunidad and obj.oportunidad.cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para editar esta gestión")
        return obj


class LlamadaDeleteView(AdminOnlyMixin, DeleteView):
    model = Llamada
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("core:llamada_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["objeto_nombre"] = f"{self.object.oportunidad.cliente.nombre_empresa} - {self.object.fecha_hora.strftime('%d/%m/%Y')}"
        context["tipo_objeto"] = "Gestión"
        return context


# Vista API para obtener contactos de una oportunidad
class OportunidadContactosAPIView(LoginRequiredMixin, View):
    def get(self, request, pk):
        try:
            oportunidad = Oportunidad.objects.get(pk=pk)
            
            # Verificar permisos
            if request.user.groups.filter(name='Ejecutivo').exists():
                if oportunidad.cliente.usuario_asignado != request.user:
                    return JsonResponse({'error': 'No tienes permiso'}, status=403)
            
            # Obtener contactos del cliente de la oportunidad
            contactos = Contacto.objects.filter(
                cliente=oportunidad.cliente, activo=True
            ).values('id', 'nombre', 'cargo')
            
            return JsonResponse({
                'success': True,
                'contactos': list(contactos),
                'cliente_id': oportunidad.cliente.id,
                'current_time': datetime.now().strftime('%Y-%m-%dT%H:%M')
            })
        except Oportunidad.DoesNotExist:
            return JsonResponse({'error': 'Oportunidad no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


# ==================== META VENTAS ====================
class MetaVentasListView(LoginRequiredMixin, TenantFilterMixin, ListView):
    model = MetaVentas
    template_name = "core/meta_ventas_list.html"
    context_object_name = "metas"
    paginate_by = 20
    tenant_field = 'empresa'


class MetaVentasCreateView(LoginRequiredMixin, CreateView):
    model = MetaVentas
    form_class = MetaVentasForm
    template_name = "core/meta_ventas_form.html"
    success_url = reverse_lazy("core:meta_ventas_list")

    def form_valid(self, form):
        try:
            form.instance.empresa = self.request.user.profile.empresa
        except Exception:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nueva Meta de Ventas"
        context["accion"] = "crear"
        return context


class MetaVentasUpdateView(LoginRequiredMixin, UpdateView):
    model = MetaVentas
    form_class = MetaVentasForm
    template_name = "core/meta_ventas_form.html"
    success_url = reverse_lazy("core:meta_ventas_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Meta de Ventas"
        context["accion"] = "editar"
        return context


# ==================== COMISIÓN ====================
class ComisionListView(LoginRequiredMixin, ComisionQuerysetFilterMixin, ListView):
    model = Comision
    template_name = "core/comision_list.html"
    context_object_name = "comisiones"
    paginate_by = 20
    login_url = 'core:login'

    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.GET.get("estado", "").strip()
        
        if estado:
            queryset = queryset.filter(estado=estado)
            
        return queryset.select_related('usuario').order_by('-periodo')


class ComisionDetailView(LoginRequiredMixin, DetailView):
    model = Comision
    template_name = "core/comision_detail.html"
    context_object_name = "comision"


# ==================== SEGUIMIENTO ====================
class SeguimientoListView(LoginRequiredMixin, SeguimientoQuerysetFilterMixin, ListView):
    model = Seguimiento
    template_name = "core/seguimiento_list.html"
    context_object_name = "seguimientos"
    paginate_by = 20
    login_url = 'core:login'

    def dispatch(self, request, *args, **kwargs):
        """Permitir acceso a Manager, Admin, Ejecutivo y Vendedor"""
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups or 'Ejecutivo' in user_groups or 'Vendedor' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.GET.get("estado", "").strip()
        tipo = self.request.GET.get("tipo", "").strip()
        
        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
            
        return queryset.select_related('oportunidad').order_by('-fecha_vencimiento')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_role'] = get_user_role(self.request.user)
        return context


class SeguimientoCreateView(LoginRequiredMixin, CreateView):
    model = Seguimiento
    form_class = SeguimientoForm
    template_name = "core/seguimiento_form.html"
    success_url = reverse_lazy("core:seguimiento_list")

    def get_initial(self):
        initial = super().get_initial()
        oportunidad_id = self.request.GET.get('oportunidad')
        if oportunidad_id:
            initial['oportunidad'] = oportunidad_id
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Nuevo Seguimiento"
        context["accion"] = "crear"
        return context


class SeguimientoDeleteView(LoginRequiredMixin, DeleteView):
    model = Seguimiento
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("core:seguimiento_list")

    def dispatch(self, request, *args, **kwargs):
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("Solo Admin o Manager pueden eliminar seguimientos.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["objeto_nombre"] = f"{self.object.oportunidad.cliente.nombre_empresa} — {self.object.tipo}"
        context["tipo_objeto"] = "Seguimiento"
        return context


class SeguimientoUpdateView(LoginRequiredMixin, UpdateView):
    """Permite editar seguimientos desde la interfaz principal"""
    model = Seguimiento
    form_class = SeguimientoForm
    template_name = "core/seguimiento_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy("core:seguimiento_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["titulo"] = "Editar Seguimiento"
        context["accion"] = "editar"
        context['user_role'] = get_user_role(self.request.user)
        return context
    
    def get_object(self, queryset=None):
        """Verifica permisos según el rol del usuario"""
        obj = super().get_object(queryset)
        user = self.request.user
        
        # Admin puede editar cualquier seguimiento
        if user.is_superuser or user.groups.filter(name='Admin').exists():
            return obj
        
        # Ejecutivo solo puede editar seguimientos de sus oportunidades
        if user.groups.filter(name='Ejecutivo').exists():
            if obj.oportunidad.cliente.usuario_asignado != user:
                raise PermissionDenied("No tienes permiso para editar este seguimiento")
        
        return obj


# ==================== ADMIN UTILITIES ====================
class FixRogersRoleView(AdminOnlyMixin, View):
    """Endpoint para cambiar el rol de un usuario a MANAGER desde el panel admin."""
    def get(self, request):
        if not request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        try:
            from django.contrib.auth import get_user_model
            from .models import UserProfile
            
            User = get_user_model()
            
            # Obtener usuario 'roy' (Rogers)
            rogers = User.objects.get(username='roy')
            
            # Obtener o crear UserProfile
            profile, created = UserProfile.objects.get_or_create(user=rogers)
            
            # Cambiar rol a MANAGER
            old_role = profile.role
            profile.role = 'MANAGER'
            profile.save()
            
            # Obtener subordinados
            subordinados = profile.subordinados.all()
            
            # Contar clientes
            clientes = Cliente.objects.filter(usuario_asignado__in=subordinados)
            
            return JsonResponse({
                'success': True,
                'message': f'Rogers\' role updated from "{old_role}" to "MANAGER"',
                'subordinados_count': subordinados.count(),
                'clientes_count': clientes.count(),
                'subordinados': [s.get_full_name() for s in subordinados]
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)


# ==================== API: COMUNAS POR REGIÓN ====================
def api_comunas(request):
    from .chile_data import REGIONES_COMUNAS
    region = request.GET.get('region', '')
    comunas = sorted(REGIONES_COMUNAS.get(region, []))
    return JsonResponse({'comunas': comunas})


# ==================== API: PROVINCIA POR COMUNA ====================
def api_provincia(request):
    from .chile_data import COMUNAS_PROVINCIA
    comuna = request.GET.get('comuna', '')
    provincia = COMUNAS_PROVINCIA.get(comuna, '')
    return JsonResponse({'provincia': provincia})


# ==================== DIAGNÓSTICO MULTI-TENANCY (solo superadmin) ====================
class DiagnosticoTenancyView(LoginRequiredMixin, View):
    """
    Vista de diagnóstico multi-tenancy. Solo accesible por superusuarios.
    URL: /admin/validar-tenancy/
    Acepta ?fix=1 para corrección automática de registros con empresa=NULL.
    """

    def get(self, request):
        if not request.user.is_superuser:
            return HttpResponseForbidden("Solo superusuarios.")

        from .models import Empresa, MetaVentas, Comision, UserProfile
        from django.contrib.auth import get_user_model
        from django.db.migrations.recorder import MigrationRecorder

        fix = request.GET.get('fix') == '1'
        User = get_user_model()
        resultados = []

        def ok(msg):
            resultados.append(('ok', msg))

        def warn(msg):
            resultados.append(('warn', msg))

        def error(msg):
            resultados.append(('error', msg))

        # 1. Migración 0010
        aplicada = MigrationRecorder.Migration.objects.filter(
            app='core', name='0010_add_empresa_to_metaventas_comision'
        ).exists()
        if aplicada:
            ok("Migración 0010 aplicada correctamente en la BD")
        else:
            error("Migración 0010 NO aplicada — ejecutar migrate")

        # 2. Empresas
        empresas = list(Empresa.objects.all())
        if empresas:
            for emp in empresas:
                ok(
                    f"Empresa [{emp.id}] «{emp.nombre}» — "
                    f"{UserProfile.objects.filter(empresa=emp).count()} usuarios, "
                    f"{Cliente.objects.filter(empresa=emp).count()} clientes, "
                    f"{MetaVentas.objects.filter(empresa=emp).count()} metas, "
                    f"{Comision.objects.filter(empresa=emp).count()} comisiones"
                )
        else:
            error("No hay empresas registradas — crear desde Django Admin")

        # 3. Perfiles sin empresa
        sin_empresa = UserProfile.objects.filter(empresa__isnull=True)
        if sin_empresa.exists():
            for p in sin_empresa:
                warn(f"Usuario sin empresa: {p.user.email} (id={p.user.id})")
        else:
            ok("Todos los usuarios tienen empresa asignada")

        # 4. Usuarios sin perfil
        sin_perfil = User.objects.filter(is_active=True, profile__isnull=True)
        if sin_perfil.exists():
            for u in sin_perfil:
                warn(f"Usuario sin UserProfile: {u.email} (id={u.id})")
        else:
            ok("Todos los usuarios activos tienen UserProfile")

        # 5. Clientes sin empresa
        clientes_null = Cliente.objects.filter(empresa__isnull=True)
        total_c = Cliente.objects.count()
        if clientes_null.exists():
            if fix:
                corregidos = self._fix_clientes(clientes_null)
                warn(f"{clientes_null.count()} clientes sin empresa → {corregidos} corregidos automáticamente (de {total_c} total)")
            else:
                warn(f"{clientes_null.count()} / {total_c} clientes sin empresa — agregar ?fix=1 para corregir")
        else:
            ok(f"Todos los clientes ({total_c}) tienen empresa")

        # 6. MetaVentas sin empresa
        metas_null = MetaVentas.objects.filter(empresa__isnull=True)
        total_m = MetaVentas.objects.count()
        if metas_null.exists():
            if fix:
                corregidos = self._fix_metas(metas_null)
                warn(f"{metas_null.count()} metas sin empresa → {corregidos} corregidas automáticamente (de {total_m} total)")
            else:
                warn(f"{metas_null.count()} / {total_m} metas sin empresa — agregar ?fix=1 para corregir")
        else:
            ok(f"Todas las metas ({total_m}) tienen empresa")

        # 7. Comisiones sin empresa
        comisiones_null = Comision.objects.filter(empresa__isnull=True)
        total_co = Comision.objects.count()
        if comisiones_null.exists():
            if fix:
                corregidos = self._fix_comisiones(comisiones_null)
                warn(f"{comisiones_null.count()} comisiones sin empresa → {corregidos} corregidas automáticamente (de {total_co} total)")
            else:
                warn(f"{comisiones_null.count()} / {total_co} comisiones sin empresa — agregar ?fix=1 para corregir")
        else:
            ok(f"Todas las comisiones ({total_co}) tienen empresa")

        estado_global = "ERROR" if any(t == 'error' for t, _ in resultados) else \
                        "ADVERTENCIAS" if any(t == 'warn' for t, _ in resultados) else "OK"

        html = self._render_html(resultados, estado_global, fix)
        return render(request, 'core/diagnostico_tenancy.html', {
            'resultados': resultados,
            'estado_global': estado_global,
            'fix': fix,
        })

    @staticmethod
    def _fix_clientes(queryset):
        corregidos = 0
        for c in queryset:
            for fuente in [c.usuario_asignado, c.creado_por]:
                if fuente:
                    try:
                        empresa = fuente.profile.empresa
                        if empresa:
                            c.empresa = empresa
                            c.save(update_fields=['empresa'])
                            corregidos += 1
                            break
                    except Exception:
                        continue
        return corregidos

    @staticmethod
    def _fix_metas(queryset):
        corregidos = 0
        for m in queryset:
            if m.usuario:
                try:
                    empresa = m.usuario.profile.empresa
                    if empresa:
                        m.empresa = empresa
                        m.save(update_fields=['empresa'])
                        corregidos += 1
                except Exception:
                    pass
        return corregidos

    @staticmethod
    def _fix_comisiones(queryset):
        corregidos = 0
        for c in queryset:
            if c.usuario:
                try:
                    empresa = c.usuario.profile.empresa
                    if empresa:
                        c.empresa = empresa
                        c.save(update_fields=['empresa'])
                        corregidos += 1
                except Exception:
                    pass
        return corregidos

    @staticmethod
    def _render_html(resultados, estado, fix):
        pass  # Se renderiza por template


# ==================== CLIENTES PROSPECTADOS ====================
class ClientesProspectadosView(LoginRequiredMixin, TemplateView):
    """Vista para Admin y Manager: muestra clientes nuevos creados manualmente por vendedores."""
    template_name = "core/clientes_prospectados.html"
    login_url = 'core:login'

    def dispatch(self, request, *args, **kwargs):
        user_groups = request.user.groups.values_list('name', flat=True)
        if request.user.is_superuser or 'Admin' in user_groups or 'Manager' in user_groups:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("No tienes permiso para acceder a esta página.")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        user_groups = user.groups.values_list('name', flat=True)
        is_manager = 'Manager' in user_groups

        # Filtrar según rol
        if is_manager:
            subordinados_ids = list(user.subordinados.values_list('user_id', flat=True))
            subordinados_ids.append(user.id)
            clientes_nuevos = Cliente.objects.filter(
                origen='MANUAL',
                usuario_asignado_id__in=subordinados_ids
            ).select_related('usuario_asignado', 'creado_por').order_by('-created_at')
        else:
            clientes_nuevos = Cliente.objects.filter(
                origen='MANUAL'
            ).select_related('usuario_asignado', 'creado_por').order_by('-created_at')

        # Filtro por vendedor opcional
        vendedor_id = self.request.GET.get('vendedor', '')
        if vendedor_id:
            clientes_nuevos = clientes_nuevos.filter(usuario_asignado_id=vendedor_id)

        # Filtro por fecha
        desde = self.request.GET.get('desde', '')
        hasta = self.request.GET.get('hasta', '')
        if desde:
            clientes_nuevos = clientes_nuevos.filter(created_at__date__gte=desde)
        if hasta:
            clientes_nuevos = clientes_nuevos.filter(created_at__date__lte=hasta)

        # Lista de vendedores disponibles para el filtro
        from django.contrib.auth import get_user_model
        User = get_user_model()
        if is_manager:
            vendedores = User.objects.filter(id__in=subordinados_ids)
        else:
            vendedores = User.objects.filter(
                groups__name__in=['Ejecutivo', 'Vendedor']
            ).distinct()

        context['clientes_nuevos'] = clientes_nuevos
        context['total'] = clientes_nuevos.count()
        context['vendedores'] = vendedores
        context['vendedor_id'] = vendedor_id
        context['desde'] = desde
        context['hasta'] = hasta
        return context


# ==================== IMPORTACIÓN DE CLIENTES ====================
class ClienteImportView(LoginRequiredMixin, View):
    """
    Importación de clientes desde Excel/CSV con preview antes de confirmar.
    - Ejecutivo: auto-asigna a sí mismo
    - Manager:   elige entre sus subordinados
    - Admin:     elige cualquier usuario
    """
    template_name = "core/cliente_import.html"
    login_url = 'core:login'

    def _get_usuarios_disponibles(self, user):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            profile = user.profile
            if profile.es_manager:
                ids = list(user.subordinados.values_list('user_id', flat=True))
                ids.append(user.id)
                return User.objects.filter(id__in=ids)
        except Exception:
            pass
        if user.is_superuser or user.groups.filter(name='Admin').exists():
            return User.objects.filter(is_active=True)
        return User.objects.filter(id=user.id)

    def _es_solo_ejecutivo(self, user):
        try:
            return user.profile.es_ejecutivo
        except Exception:
            return user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists()

    def _parsear_archivo(self, archivo):
        """Lee Excel o CSV y retorna lista de dicts con los datos."""
        nombre = archivo.name.lower()
        filas = []
        errores_parse = []

        try:
            if nombre.endswith('.xlsx') or nombre.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
                ws = wb.active
                # Normaliza headers: minúsculas, espacios y guiones → guión bajo
                headers = [
                    str(c.value).strip().lower().replace(' ', '_').replace('-', '_') if c.value else ''
                    for c in next(ws.iter_rows(min_row=1, max_row=1))
                ]
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if all(v is None for v in row):
                        continue
                    filas.append({'_fila': i, **{headers[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row)}})
            elif nombre.endswith('.csv'):
                import csv
                content = archivo.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for i, row in enumerate(reader, start=2):
                    filas.append({'_fila': i, **{
                        k.strip().lower().replace(' ', '_').replace('-', '_'): v.strip()
                        for k, v in row.items()
                    }})
            else:
                errores_parse.append('Formato no soportado. Usa .xlsx o .csv')
        except Exception as e:
            errores_parse.append(f'Error al leer el archivo: {str(e)}')

        return filas, errores_parse

    def _procesar_preview(self, filas, usuario_asignado):
        """Analiza las filas y clasifica cada una como NUEVO, ACTUALIZAR o ERROR."""
        from .models import Cliente
        preview = []
        for fila in filas:
            rut = fila.get('rut', '').replace('.', '').strip()
            nombre = (fila.get('nombre_empresa') or fila.get('nombre') or
                      fila.get('razon_social') or fila.get('empresa') or '').strip()
            num_fila = fila.get('_fila', '?')

            if not rut:
                preview.append({'fila': num_fila, 'rut': '-', 'nombre': nombre or '-',
                                 'estado': 'ERROR', 'detalle': 'RUT vacío'})
                continue
            if not nombre:
                preview.append({'fila': num_fila, 'rut': rut, 'nombre': '-',
                                 'estado': 'ERROR', 'detalle': 'Nombre de empresa vacío'})
                continue

            existe = Cliente.objects.filter(rut=rut).first()
            if existe:
                preview.append({'fila': num_fila, 'rut': rut, 'nombre': nombre,
                                 'estado': 'ACTUALIZAR', 'detalle': f'Ya existe (id={existe.pk})'})
            else:
                preview.append({'fila': num_fila, 'rut': rut, 'nombre': nombre,
                                 'estado': 'NUEVO', 'detalle': 'Se creará'})
        return preview

    def _ejecutar_import(self, filas, usuario_asignado, reasignar=False, importado_por=None):
        """Importa las filas válidas. Retorna estadísticas."""
        from .models import Cliente
        from django.contrib.auth import get_user_model
        User = get_user_model()

        creados = 0
        actualizados = 0
        errores = []
        filas_incompletas = []  # filas sin nombre de empresa
        errores_tecnicos = []   # errores de base de datos

        CAMPO_MAP = {
            'nombre_empresa': ['nombre_empresa', 'nombre', 'razon_social', 'empresa'],
            'sector':         ['sector', 'rubro'],
            'tipo_cliente':   ['tipo_cliente', 'tipo'],
            'estado':         ['estado'],
            'segmento':       ['segmento'],
            'subrango':       ['subrango'],
            'comuna':         ['comuna'],
            'region':         ['region', 'región'],
            'renta_uf_total': ['renta_uf_total', 'renta_uf'],
            'q_total_lineas': ['q_total_lineas', 'lineas', 'q_lineas'],
            'bam':            ['bam'],
            'm2m':            ['m2m'],
            'voz':            ['voz'],
            'edv':            ['edv'],
        }

        for fila in filas:
            rut = fila.get('rut', '').replace('.', '').strip()
            if not rut:
                continue

            datos = {'rut': rut, 'usuario_asignado': usuario_asignado}
            for campo, aliases in CAMPO_MAP.items():
                for alias in aliases:
                    val = fila.get(alias, '')
                    if val:
                        datos[campo] = val
                        break

            if not datos.get('nombre_empresa'):
                filas_incompletas.append(fila.get('_fila', '?'))
                continue

            # Si el cliente ya existe y no se pidió reasignar, conservar usuario_asignado actual
            defaults = {k: v for k, v in datos.items() if k != 'rut'}
            if not reasignar:
                existente = Cliente.objects.filter(rut=rut).first()
                if existente:
                    defaults.pop('usuario_asignado', None)

            # Origen: ejecutivo/vendedor -> MANUAL (prospectado), admin/manager -> IMPORTADO
            if not existente:
                es_ejecutivo_importando = (
                    importado_por and
                    not importado_por.is_superuser and
                    not importado_por.groups.filter(name__in=['Admin', 'Manager']).exists()
                )
                defaults['origen'] = 'MANUAL' if es_ejecutivo_importando else 'IMPORTADO'
                if es_ejecutivo_importando:
                    defaults['creado_por'] = importado_por

            # Auto-asignar empresa del usuario que importa (Capa 1 anti-fuga)
            if importado_por and not importado_por.is_superuser:
                try:
                    empresa_importador = importado_por.profile.empresa
                    if empresa_importador:
                        defaults['empresa'] = empresa_importador
                except Exception:
                    pass

            # Reintentos para manejar bloqueos de SQLite
            for intento in range(3):
                try:
                    cliente, creado = Cliente.objects.update_or_create(
                        rut=rut,
                        defaults=defaults
                    )
                    if creado:
                        creados += 1
                    else:
                        actualizados += 1
                    break
                except Exception:
                    import time as _time
                    _time.sleep(0.1)
                    if intento == 2:
                        errores_tecnicos.append(rut)

        return {
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores,
            'filas_incompletas': filas_incompletas,
            'errores_tecnicos': errores_tecnicos,
        }

    def get(self, request):
        usuarios = self._get_usuarios_disponibles(request.user)
        es_ejecutivo = self._es_solo_ejecutivo(request.user)
        return render(request, self.template_name, {
            'step': 'upload',
            'usuarios': usuarios,
            'es_ejecutivo': es_ejecutivo,
        })

    def post(self, request):
        usuarios = self._get_usuarios_disponibles(request.user)
        es_ejecutivo = self._es_solo_ejecutivo(request.user)

        # PASO 2 - Confirmar e importar
        if request.POST.get('action') == 'confirmar':
            import json as _json
            filas_json = request.session.get('import_filas')
            usuario_id = request.session.get('import_usuario_id')

            if not filas_json or not usuario_id:
                return redirect('core:cliente_import')

            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                usuario_asignado = User.objects.get(pk=usuario_id)
            except User.DoesNotExist:
                usuario_asignado = request.user

            reasignar = request.session.get('import_reasignar', False)
            filas = _json.loads(filas_json)
            resultado = self._ejecutar_import(filas, usuario_asignado, reasignar=reasignar, importado_por=request.user)

            # Limpiar sesión
            del request.session['import_filas']
            del request.session['import_usuario_id']
            request.session.pop('import_reasignar', None)

            filas_inc = resultado.get('filas_incompletas', [])
            err_tec   = resultado.get('errores_tecnicos', [])

            return render(request, self.template_name, {
                'step': 'resultado',
                'creados': resultado['creados'],
                'actualizados': resultado['actualizados'],
                'errores': resultado['errores'],
                'filas_incompletas': filas_inc,
                'errores_tecnicos': err_tec,
                'total': resultado['creados'] + resultado['actualizados'] + len(filas_inc) + len(err_tec),
            })

        # PASO 1 - Subir archivo y mostrar preview
        archivo = request.FILES.get('archivo')
        if not archivo:
            return render(request, self.template_name, {
                'step': 'upload', 'usuarios': usuarios, 'es_ejecutivo': es_ejecutivo,
                'error': 'Debes seleccionar un archivo.'
            })

        # Determinar usuario asignado
        if es_ejecutivo:
            usuario_asignado = request.user
        else:
            usuario_id = request.POST.get('usuario_asignado')
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                usuario_asignado = User.objects.get(pk=usuario_id)
            except (User.DoesNotExist, ValueError, TypeError):
                usuario_asignado = request.user

        filas, errores_parse = self._parsear_archivo(archivo)
        if errores_parse:
            return render(request, self.template_name, {
                'step': 'upload', 'usuarios': usuarios, 'es_ejecutivo': es_ejecutivo,
                'error': ' / '.join(errores_parse)
            })

        preview = self._procesar_preview(filas, usuario_asignado)

        # Guardar en sesión para el paso 2
        import json as _json
        request.session['import_filas'] = _json.dumps(filas)
        request.session['import_usuario_id'] = usuario_asignado.pk
        request.session['import_reasignar'] = request.POST.get('reasignar_existentes') == 'on'

        nuevos = sum(1 for p in preview if p['estado'] == 'NUEVO')
        actualizar = sum(1 for p in preview if p['estado'] == 'ACTUALIZAR')
        con_error = sum(1 for p in preview if p['estado'] == 'ERROR')

        return render(request, self.template_name, {
            'step': 'preview',
            'preview': preview,
            'usuario_asignado': usuario_asignado,
            'reasignar': request.session['import_reasignar'],
            'nuevos': nuevos,
            'actualizar': actualizar,
            'con_error': con_error,
            'total': len(preview),
        })


# ==================== SINDICATO MVP (DÍA 4) ====================

def _grupo_nombre_normalizado(nombre):
    return (
        (nombre or '')
        .strip()
        .lower()
        .replace('á', 'a')
        .replace('é', 'e')
        .replace('í', 'i')
        .replace('ó', 'o')
        .replace('ú', 'u')
    )


def _roles_usuario(user):
    return {_grupo_nombre_normalizado(n) for n in user.groups.values_list('name', flat=True)}


def _es_admin_sindicato(user):
    roles = _roles_usuario(user)
    return user.is_superuser or 'admin' in roles or 'administracion' in roles


def _es_tesoreria_sindicato(user):
    roles = _roles_usuario(user)
    return user.is_superuser or 'tesoreria' in roles


def _es_dirigente_sindicato(user):
    roles = _roles_usuario(user)
    return user.is_superuser or 'dirigente' in roles


def _rut_normalizado_simple(rut):
    if not rut:
        return ''
    limpio = ''.join(ch for ch in str(rut).upper() if ch.isdigit() or ch == 'K')
    if len(limpio) < 2:
        return limpio
    return f"{limpio[:-1]}-{limpio[-1]}"


PERIODO_REGEX_SIND = re.compile(r'^\d{4}-(0[1-9]|1[0-2])$')


def _normalizar_rut_import(rut):
    if not rut:
        return ''
    limpio = re.sub(r'[^0-9kK]', '', str(rut)).upper()
    if len(limpio) < 2:
        return limpio
    return f"{limpio[:-1]}-{limpio[-1]}"


def _validar_rut_chileno_import(rut_normalizado):
    if not rut_normalizado or '-' not in rut_normalizado:
        return False
    cuerpo, dv = rut_normalizado.split('-', 1)
    if not cuerpo.isdigit() or not dv:
        return False

    dv = dv.upper()
    suma = 0
    multiplo = 2
    for digito in reversed(cuerpo):
        suma += int(digito) * multiplo
        multiplo = multiplo + 1 if multiplo < 7 else 2

    resto = 11 - (suma % 11)
    dv_esperado = '0' if resto == 11 else 'K' if resto == 10 else str(resto)
    return dv == dv_esperado


class SindicatoTenantMixin(LoginRequiredMixin):
    login_url = 'core:login'

    def get_empresa_usuario(self):
        user = self.request.user
        if user.is_superuser:
            empresa_id = self.request.GET.get('empresa_id') or self.request.POST.get('empresa_id')
            if empresa_id:
                try:
                    return Empresa.objects.get(pk=empresa_id)
                except Empresa.DoesNotExist:
                    raise Http404('Empresa no encontrada')
            return None

        try:
            return user.profile.empresa
        except Exception:
            return None

    def filtrar_por_tenant(self, queryset):
        empresa = self.get_empresa_usuario()
        if self.request.user.is_superuser and empresa is None:
            return queryset
        if empresa is None:
            return queryset.none()
        return queryset.filter(empresa=empresa)


class SindicatoRolePermissionMixin:
    permiso_ver = None
    permiso_editar = None

    def _check_permiso(self, permiso):
        user = self.request.user
        if user.is_superuser:
            return True
        if permiso == 'socios_beneficios_ver':
            return _es_admin_sindicato(user) or _es_dirigente_sindicato(user)
        if permiso == 'socios_beneficios_editar':
            return _es_admin_sindicato(user)
        if permiso == 'movimientos_ver':
            return _es_admin_sindicato(user) or _es_tesoreria_sindicato(user) or _es_dirigente_sindicato(user)
        if permiso == 'movimientos_editar':
            return _es_tesoreria_sindicato(user)
        if permiso == 'movimientos_importar':
            return _es_admin_sindicato(user) or _es_tesoreria_sindicato(user)
        if permiso == 'consolidado_ver':
            return _es_admin_sindicato(user) or _es_tesoreria_sindicato(user) or _es_dirigente_sindicato(user)
        if permiso == 'consolidado_editar':
            return _es_admin_sindicato(user) or _es_tesoreria_sindicato(user)
        if permiso == 'consulta_rut':
            return _es_admin_sindicato(user) or _es_tesoreria_sindicato(user) or _es_dirigente_sindicato(user)
        return False

    def dispatch(self, request, *args, **kwargs):
        permiso = self.permiso_ver
        if request.method in ('POST', 'PUT', 'PATCH', 'DELETE'):
            permiso = self.permiso_editar or self.permiso_ver
        if not self._check_permiso(permiso):
            return HttpResponseForbidden('No tienes permiso para acceder a esta sección.')
        return super().dispatch(request, *args, **kwargs)


class SocioSindicatoListView(SindicatoTenantMixin, SindicatoRolePermissionMixin, ListView):
    model = SocioSindicato
    template_name = 'core/sindicato/socio_list.html'
    context_object_name = 'socios'
    permiso_ver = 'socios_beneficios_ver'

    def get_queryset(self):
        qs = self.filtrar_por_tenant(SocioSindicato.objects.all())
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(rut__icontains=q) | Q(nombre__icontains=q))
        return qs.order_by('nombre')


class SocioSindicatoCreateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, CreateView):
    model = SocioSindicato
    form_class = SocioSindicatoForm
    template_name = 'core/sindicato/socio_form.html'
    success_url = reverse_lazy('core:sindicato_socio_list')
    permiso_ver = 'socios_beneficios_editar'
    permiso_editar = 'socios_beneficios_editar'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs

    def form_valid(self, form):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            form.add_error(None, 'No se pudo determinar la empresa del usuario.')
            return self.form_invalid(form)
        form.instance.empresa = empresa
        return super().form_valid(form)


class SocioSindicatoUpdateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, UpdateView):
    model = SocioSindicato
    form_class = SocioSindicatoForm
    template_name = 'core/sindicato/socio_form.html'
    success_url = reverse_lazy('core:sindicato_socio_list')
    permiso_ver = 'socios_beneficios_editar'
    permiso_editar = 'socios_beneficios_editar'

    def get_queryset(self):
        return self.filtrar_por_tenant(SocioSindicato.objects.all())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs


class TipoBeneficioSindicatoListView(SindicatoTenantMixin, SindicatoRolePermissionMixin, ListView):
    model = TipoBeneficioSindicato
    template_name = 'core/sindicato/beneficio_list.html'
    context_object_name = 'beneficios'
    permiso_ver = 'socios_beneficios_ver'

    def get_queryset(self):
        return self.filtrar_por_tenant(TipoBeneficioSindicato.objects.all()).order_by('orden_export', 'nombre')


class TipoBeneficioSindicatoCreateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, CreateView):
    model = TipoBeneficioSindicato
    form_class = TipoBeneficioSindicatoForm
    template_name = 'core/sindicato/beneficio_form.html'
    success_url = reverse_lazy('core:sindicato_beneficio_list')
    permiso_ver = 'socios_beneficios_editar'
    permiso_editar = 'socios_beneficios_editar'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs

    def form_valid(self, form):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            form.add_error(None, 'No se pudo determinar la empresa del usuario.')
            return self.form_invalid(form)
        form.instance.empresa = empresa
        return super().form_valid(form)


class TipoBeneficioSindicatoUpdateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, UpdateView):
    model = TipoBeneficioSindicato
    form_class = TipoBeneficioSindicatoForm
    template_name = 'core/sindicato/beneficio_form.html'
    success_url = reverse_lazy('core:sindicato_beneficio_list')
    permiso_ver = 'socios_beneficios_editar'
    permiso_editar = 'socios_beneficios_editar'

    def get_queryset(self):
        return self.filtrar_por_tenant(TipoBeneficioSindicato.objects.all())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs


class MovimientoSindicatoListView(SindicatoTenantMixin, SindicatoRolePermissionMixin, ListView):
    model = MovimientoSindicato
    template_name = 'core/sindicato/movimiento_list.html'
    context_object_name = 'movimientos'
    permiso_ver = 'movimientos_ver'

    def get_queryset(self):
        qs = self.filtrar_por_tenant(
            MovimientoSindicato.objects.select_related('socio', 'tipo_beneficio')
        )
        periodo = self.request.GET.get('periodo', '').strip()
        if periodo:
            qs = qs.filter(periodo=periodo)
        return qs.order_by('-periodo', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['puede_importar'] = self._check_permiso('movimientos_importar')
        return context


class MovimientoSindicatoCreateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, CreateView):
    model = MovimientoSindicato
    form_class = MovimientoSindicatoForm
    template_name = 'core/sindicato/movimiento_form.html'
    success_url = reverse_lazy('core:sindicato_movimiento_list')
    permiso_ver = 'movimientos_editar'
    permiso_editar = 'movimientos_editar'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warnings'] = getattr(self.get_form(), 'warnings', [])
        return context

    def form_valid(self, form):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            form.add_error(None, 'No se pudo determinar la empresa del usuario.')
            return self.form_invalid(form)
        form.instance.empresa = empresa
        form.instance.creado_por = self.request.user
        return super().form_valid(form)


class MovimientoSindicatoUpdateView(SindicatoTenantMixin, SindicatoRolePermissionMixin, UpdateView):
    model = MovimientoSindicato
    form_class = MovimientoSindicatoForm
    template_name = 'core/sindicato/movimiento_form.html'
    success_url = reverse_lazy('core:sindicato_movimiento_list')
    permiso_ver = 'movimientos_editar'
    permiso_editar = 'movimientos_editar'

    def get_queryset(self):
        return self.filtrar_por_tenant(MovimientoSindicato.objects.all())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['empresa'] = self.get_empresa_usuario()
        return kwargs

    def form_valid(self, form):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            form.add_error(None, 'No se pudo determinar la empresa del usuario.')
            return self.form_invalid(form)
        form.instance.empresa = empresa
        return super().form_valid(form)


class MovimientoSindicatoImportView(SindicatoTenantMixin, SindicatoRolePermissionMixin, TemplateView):
    template_name = 'core/sindicato/movimiento_import.html'
    permiso_ver = 'movimientos_importar'
    permiso_editar = 'movimientos_importar'
    session_rows_key = 'sindicato_import_rows'
    session_meta_key = 'sindicato_import_meta'

    def _normalizar_header(self, text):
        value = (str(text or '').strip().lower())
        value = (
            value.replace('á', 'a')
            .replace('é', 'e')
            .replace('í', 'i')
            .replace('ó', 'o')
            .replace('ú', 'u')
            .replace('ñ', 'n')
        )
        return value.replace(' ', '_').replace('-', '_')

    def _resolver_valor(self, row, aliases):
        for key in aliases:
            val = row.get(key)
            if val is not None and str(val).strip() != '':
                return str(val).strip()
        return ''

    def _leer_archivo(self, archivo):
        name = (archivo.name or '').lower()
        rows = []
        errors = []

        try:
            if name.endswith('.xlsx') or name.endswith('.xls'):
                import openpyxl

                content = archivo.read()
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                ws = wb.active
                headers = [self._normalizar_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for idx, line in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if all(v is None or str(v).strip() == '' for v in line):
                        continue
                    item = {'_fila': idx}
                    for col, raw in enumerate(line):
                        if col >= len(headers):
                            continue
                        key = headers[col]
                        if key:
                            item[key] = '' if raw is None else str(raw).strip()
                    rows.append(item)
            elif name.endswith('.csv'):
                content = archivo.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                for idx, line in enumerate(reader, start=2):
                    item = {'_fila': idx}
                    for k, v in (line or {}).items():
                        nk = self._normalizar_header(k)
                        if nk:
                            item[nk] = '' if v is None else str(v).strip()
                    if any(str(v).strip() for k, v in item.items() if k != '_fila'):
                        rows.append(item)
            else:
                errors.append('Formato no soportado. Usa archivo .xlsx, .xls o .csv.')
        except Exception as exc:
            errors.append(f'No fue posible leer el archivo: {exc}')

        return rows, errors

    def _parsear_monto(self, raw):
        val = (raw or '').strip()
        if not val:
            raise InvalidOperation('Monto vacío')
        normalized = val.replace('.', '').replace(',', '.')
        return Decimal(normalized)

    def _beneficio_empresa(self, empresa, beneficio_id):
        try:
            return TipoBeneficioSindicato.objects.get(pk=beneficio_id, empresa=empresa)
        except TipoBeneficioSindicato.DoesNotExist:
            return None

    def _periodo_abierto(self, empresa, periodo):
        return not ConsolidadoMensualSindicato.objects.filter(
            empresa=empresa,
            periodo=periodo,
            estado__in=[
                ConsolidadoMensualSindicato.ESTADO_CERRADO,
                ConsolidadoMensualSindicato.ESTADO_EXPORTADO,
            ],
        ).exists()

    def _validar_y_previsualizar(self, empresa, beneficio, periodo, filas, source_tag):
        aliases = {
            'rut': ['rut'],
            'nombre': ['nombre', 'socio', 'nombre_socio'],
            'monto': ['monto', 'valor'],
            'observacion': ['observacion', 'observaciones'],
            'referencia_externa': ['referencia_externa', 'referencia', 'folio'],
            'site': ['site'],
            'cantidad': ['cantidad'],
            'tipo_vale': ['tipo_de_vale', 'tipo_vale'],
            'numero_linea': ['numero_linea', 'numero_de_linea'],
            'fecha_entrega': ['fecha_entrega'],
        }
        preview = []
        valid_rows = []
        refs_seen = set()

        for row in filas:
            fila_num = row.get('_fila', '?')
            rut_raw = self._resolver_valor(row, aliases['rut'])
            nombre = self._resolver_valor(row, aliases['nombre'])
            monto_raw = self._resolver_valor(row, aliases['monto'])
            observacion = self._resolver_valor(row, aliases['observacion'])
            referencia = self._resolver_valor(row, aliases['referencia_externa'])

            extras = []
            for field, title in (
                ('site', 'Site'),
                ('cantidad', 'Cantidad'),
                ('tipo_vale', 'Tipo vale'),
                ('numero_linea', 'Numero linea'),
                ('fecha_entrega', 'Fecha entrega'),
            ):
                value = self._resolver_valor(row, aliases[field])
                if value:
                    extras.append(f'{title}: {value}')

            if extras:
                observacion = f"{observacion} | {' | '.join(extras)}".strip(' |')

            rut = _normalizar_rut_import(rut_raw)
            if not rut or not _validar_rut_chileno_import(rut):
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut_raw or '-',
                        'nombre': nombre or '-',
                        'monto': monto_raw or '-',
                        'estado': 'RECHAZADA',
                        'motivo': 'RUT inválido o vacío.',
                    }
                )
                continue

            if not nombre:
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut,
                        'nombre': '-',
                        'monto': monto_raw or '-',
                        'estado': 'RECHAZADA',
                        'motivo': 'Nombre es obligatorio.',
                    }
                )
                continue

            try:
                monto = self._parsear_monto(monto_raw)
            except InvalidOperation:
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut,
                        'nombre': nombre,
                        'monto': monto_raw or '-',
                        'estado': 'RECHAZADA',
                        'motivo': 'Monto inválido.',
                    }
                )
                continue

            if monto <= 0:
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut,
                        'nombre': nombre,
                        'monto': str(monto),
                        'estado': 'RECHAZADA',
                        'motivo': 'Monto debe ser mayor a cero.',
                    }
                )
                continue

            socio = SocioSindicato.objects.filter(empresa=empresa, rut=rut).first()
            if socio is None:
                estado_socio = 'CREAR_SOCIO'
            else:
                estado_socio = 'SOCIO_EXISTENTE'

            referencia_informada = bool(referencia)
            if not referencia:
                referencia = f"AUTO-{source_tag}-{fila_num}"

            dup_qs = MovimientoSindicato.objects.filter(
                empresa=empresa,
                socio=socio,
                tipo_beneficio=beneficio,
                periodo=periodo,
                referencia_externa=referencia,
            ) if socio else MovimientoSindicato.objects.none()

            if referencia_informada and dup_qs.exists():
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut,
                        'nombre': nombre,
                        'monto': str(monto),
                        'estado': 'RECHAZADA',
                        'motivo': 'Referencia externa duplicada para este beneficio y período.',
                    }
                )
                continue

            key_ref = (rut, referencia)
            if key_ref in refs_seen:
                preview.append(
                    {
                        'fila': fila_num,
                        'rut': rut,
                        'nombre': nombre,
                        'monto': str(monto),
                        'estado': 'RECHAZADA',
                        'motivo': 'Referencia duplicada dentro del archivo.',
                    }
                )
                continue
            refs_seen.add(key_ref)

            valid_rows.append(
                {
                    'fila': fila_num,
                    'rut': rut,
                    'nombre': nombre,
                    'monto': str(monto.quantize(Decimal('1'))),
                    'observacion': observacion,
                    'referencia_externa': referencia,
                    'referencia_informada': referencia_informada,
                }
            )
            preview.append(
                {
                    'fila': fila_num,
                    'rut': rut,
                    'nombre': nombre,
                    'monto': str(monto.quantize(Decimal('1'))),
                    'estado': 'VALIDA',
                    'motivo': 'OK' if estado_socio == 'SOCIO_EXISTENTE' else 'Se creará socio mínimo.',
                }
            )

        return preview, valid_rows

    def _ejecutar_importacion(self, empresa, beneficio, periodo, rows):
        creados = 0
        rechazados = []

        for row in rows:
            socio, _ = SocioSindicato.objects.get_or_create(
                empresa=empresa,
                rut=row['rut'],
                defaults={
                    'nombre': row['nombre'],
                    'site': '',
                    'estado_laboral': SocioSindicato.ESTADO_LABORAL_ACTIVO,
                    'estado': SocioSindicato.ESTADO_ACTIVO,
                },
            )

            if socio.empresa_id != empresa.id:
                rechazados.append({'fila': row['fila'], 'motivo': 'Socio pertenece a otro tenant.'})
                continue

            ref = (row.get('referencia_externa') or '').strip()
            if MovimientoSindicato.objects.filter(
                empresa=empresa,
                socio=socio,
                tipo_beneficio=beneficio,
                periodo=periodo,
                referencia_externa=ref,
            ).exists():
                rechazados.append({'fila': row['fila'], 'motivo': 'Duplicado por referencia externa.'})
                continue

            MovimientoSindicato.objects.create(
                empresa=empresa,
                socio=socio,
                tipo_beneficio=beneficio,
                periodo=periodo,
                monto=Decimal(row['monto']),
                estado=MovimientoSindicato.ESTADO_PENDIENTE,
                observacion=row.get('observacion') or '',
                referencia_externa=ref,
                creado_por=self.request.user,
            )
            creados += 1

        return creados, rechazados

    def _build_context(self, empresa, **extra):
        beneficios = TipoBeneficioSindicato.objects.filter(
            empresa=empresa,
            estado=TipoBeneficioSindicato.ESTADO_ACTIVO,
        ).order_by('orden_export', 'nombre') if empresa else TipoBeneficioSindicato.objects.none()
        context = {
            'step': 'upload',
            'beneficios': beneficios,
        }
        context.update(extra)
        return context

    def get(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            return render(
                request,
                self.template_name,
                self._build_context(None, error='No se pudo determinar la empresa del usuario.'),
            )
        return render(request, self.template_name, self._build_context(empresa))

    def post(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        if empresa is None:
            return render(
                request,
                self.template_name,
                self._build_context(None, error='No se pudo determinar la empresa del usuario.'),
            )

        action = request.POST.get('action') or 'preview'

        if action == 'confirmar':
            payload_rows = request.session.get(self.session_rows_key)
            payload_meta = request.session.get(self.session_meta_key)
            if not payload_rows or not payload_meta:
                return redirect('core:sindicato_movimiento_import')

            try:
                rows = json.loads(payload_rows)
                meta = json.loads(payload_meta)
            except Exception:
                return redirect('core:sindicato_movimiento_import')

            beneficio = self._beneficio_empresa(empresa, meta.get('beneficio_id'))
            periodo = (meta.get('periodo') or '').strip()
            if beneficio is None:
                return render(
                    request,
                    self.template_name,
                    self._build_context(empresa, error='Beneficio inválido o de otra empresa.'),
                )
            if beneficio.estado != TipoBeneficioSindicato.ESTADO_ACTIVO:
                return render(
                    request,
                    self.template_name,
                    self._build_context(empresa, error='No se puede importar sobre un beneficio inactivo.'),
                )
            if not PERIODO_REGEX_SIND.match(periodo) or not self._periodo_abierto(empresa, periodo):
                return render(
                    request,
                    self.template_name,
                    self._build_context(empresa, error='El período está cerrado o es inválido.'),
                )

            creados, rechazos_extra = self._ejecutar_importacion(empresa, beneficio, periodo, rows)
            preview = json.loads(meta.get('preview') or '[]')
            rechazadas_preview = [r for r in preview if r.get('estado') == 'RECHAZADA']
            rechazadas = len(rechazadas_preview) + len(rechazos_extra)

            total_filas = int(meta.get('total', 0))
            nombre_archivo = (meta.get('nombre_archivo') or '').strip()
            AuditoriaSindicato.objects.create(
                empresa=empresa,
                usuario=request.user,
                accion='IMPORTAR_MOVIMIENTOS',
                entidad='MovimientoSindicato',
                entidad_id=f'{beneficio.id}:{periodo}'[:40],
                periodo=periodo,
                resumen=(
                    f'Importación movimientos {periodo} | beneficio {beneficio.codigo} '
                    f'| leídas={total_filas} | importadas={creados} | rechazadas={rechazadas}'
                )[:255],
                payload={
                    'total_filas_leidas': total_filas,
                    'total_importadas': creados,
                    'total_rechazadas': rechazadas,
                    'tipo_beneficio_id': beneficio.id,
                    'tipo_beneficio_codigo': beneficio.codigo,
                    'tipo_beneficio_nombre': beneficio.nombre,
                    'periodo': periodo,
                    'nombre_archivo': nombre_archivo,
                },
            )

            request.session.pop(self.session_rows_key, None)
            request.session.pop(self.session_meta_key, None)

            return render(
                request,
                self.template_name,
                self._build_context(
                    empresa,
                    step='resultado',
                    total_filas=meta.get('total', 0),
                    validas=meta.get('validas', 0),
                    rechazadas=rechazadas,
                    movimientos_creados=creados,
                    preview=preview,
                    rechazos_confirmacion=rechazos_extra,
                    beneficio=beneficio,
                    periodo=periodo,
                ),
            )

        beneficio_id = request.POST.get('tipo_beneficio')
        periodo = (request.POST.get('periodo') or '').strip()
        archivo = request.FILES.get('archivo')

        beneficio = self._beneficio_empresa(empresa, beneficio_id)
        if beneficio is None:
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error='Selecciona un tipo de beneficio válido.'),
            )
        if beneficio.estado != TipoBeneficioSindicato.ESTADO_ACTIVO:
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error='El beneficio seleccionado está inactivo.'),
            )
        if not PERIODO_REGEX_SIND.match(periodo):
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error='Periodo inválido. Usa formato YYYY-MM.'),
            )
        if not self._periodo_abierto(empresa, periodo):
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error='El período está cerrado para nuevos movimientos.'),
            )
        if not archivo:
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error='Debes subir un archivo para previsualizar.'),
            )

        raw_rows, parse_errors = self._leer_archivo(archivo)
        if parse_errors:
            return render(
                request,
                self.template_name,
                self._build_context(empresa, error=' / '.join(parse_errors)),
            )

        source_tag = hashlib.sha1(f'{archivo.name}|{periodo}|{beneficio.id}'.encode('utf-8')).hexdigest()[:12]
        preview, valid_rows = self._validar_y_previsualizar(empresa, beneficio, periodo, raw_rows, source_tag)

        request.session[self.session_rows_key] = json.dumps(valid_rows)
        request.session[self.session_meta_key] = json.dumps(
            {
                'beneficio_id': beneficio.id,
                'periodo': periodo,
                'total': len(preview),
                'validas': len(valid_rows),
                'preview': json.dumps(preview),
                'nombre_archivo': archivo.name,
            }
        )

        return render(
            request,
            self.template_name,
            self._build_context(
                empresa,
                step='preview',
                total_filas=len(preview),
                validas=len(valid_rows),
                rechazadas=len(preview) - len(valid_rows),
                preview=preview,
                beneficio=beneficio,
                periodo=periodo,
            ),
        )


class ConsultaRutSindicatoView(SindicatoTenantMixin, SindicatoRolePermissionMixin, TemplateView):
    template_name = 'core/sindicato/consulta_rut.html'
    permiso_ver = 'consulta_rut'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa = self.get_empresa_usuario()
        rut_query = self.request.GET.get('rut', '').strip()
        rut = _rut_normalizado_simple(rut_query)
        periodo_actual = datetime.now().strftime('%Y-%m')

        context.update(
            {
                'rut_buscado': rut_query,
                'periodo_actual': periodo_actual,
                'socio': None,
                'total_periodo': 0,
                'montos_por_beneficio': [],
                'observaciones': [],
                'historial_periodos': [],
            }
        )

        if not rut:
            return context

        socios_qs = SocioSindicato.objects.all()
        if self.request.user.is_superuser and empresa is None:
            socio = socios_qs.filter(rut=rut).select_related('empresa').first()
        else:
            if empresa is None:
                return context
            socio = socios_qs.filter(empresa=empresa, rut=rut).select_related('empresa').first()

        if not socio:
            return context

        movs = MovimientoSindicato.objects.filter(empresa=socio.empresa, socio=socio)
        movs_periodo = movs.filter(periodo=periodo_actual).exclude(estado=MovimientoSindicato.ESTADO_RECHAZADO)

        context['socio'] = socio
        context['total_periodo'] = movs_periodo.aggregate(total=Sum('monto'))['total'] or 0
        context['montos_por_beneficio'] = list(
            movs_periodo.values('tipo_beneficio__nombre').annotate(total=Sum('monto')).order_by('tipo_beneficio__nombre')
        )
        context['observaciones'] = list(
            movs_periodo.exclude(observacion__isnull=True).exclude(observacion='').values(
                'periodo', 'tipo_beneficio__nombre', 'observacion', 'monto'
            )[:20]
        )
        context['historial_periodos'] = list(
            movs.exclude(estado=MovimientoSindicato.ESTADO_RECHAZADO)
            .values('periodo')
            .annotate(total=Sum('monto'))
            .order_by('-periodo')[:6]
        )
        return context


class ConsolidadoSindicatoHistorialView(SindicatoTenantMixin, SindicatoRolePermissionMixin, ListView):
    model = ConsolidadoMensualSindicato
    template_name = 'core/sindicato/consolidado_historial.html'
    context_object_name = 'consolidados'
    permiso_ver = 'consolidado_ver'

    def get_queryset(self):
        qs = self.filtrar_por_tenant(ConsolidadoMensualSindicato.objects.all())
        periodo = self.request.GET.get('periodo', '').strip()
        if periodo:
            qs = qs.filter(periodo=periodo)
        return qs.order_by('-periodo', '-updated_at')


class ConsolidadoSindicatoDetalleView(SindicatoTenantMixin, SindicatoRolePermissionMixin, DetailView):
    model = ConsolidadoMensualSindicato
    template_name = 'core/sindicato/consolidado_detalle.html'
    context_object_name = 'consolidado'
    permiso_ver = 'consolidado_ver'

    def get_queryset(self):
        return self.filtrar_por_tenant(ConsolidadoMensualSindicato.objects.all())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        consolidado = self.object
        context['detalles'] = (
            ConsolidadoDetalleSindicato.objects.select_related('socio', 'tipo_beneficio')
            .filter(consolidado=consolidado)
            .order_by('socio__nombre', 'tipo_beneficio__orden_export', 'tipo_beneficio__nombre')
        )
        context['puede_editar_consolidado'] = self._check_permiso('consolidado_editar')
        return context


class ConsolidadoSindicatoGenerarView(SindicatoTenantMixin, SindicatoRolePermissionMixin, View):
    permiso_ver = 'consolidado_editar'
    permiso_editar = 'consolidado_editar'

    def post(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        periodo = (request.POST.get('periodo') or '').strip()
        if empresa is None:
            return HttpResponseForbidden('No se pudo determinar la empresa del usuario.')

        try:
            resultado = generar_o_recalcular_consolidado(empresa=empresa, periodo=periodo, usuario=request.user)
            messages.success(
                request,
                f"{resultado.accion} OK para {resultado.periodo}. Detalles: {resultado.total_detalles_generados}.",
            )
        except (ValueError, ConsolidadoBloqueadoError) as exc:
            messages.error(request, str(exc))
        return redirect('core:sindicato_consolidado_historial')


class ConsolidadoSindicatoRecalcularView(SindicatoTenantMixin, SindicatoRolePermissionMixin, View):
    permiso_ver = 'consolidado_editar'
    permiso_editar = 'consolidado_editar'

    def post(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        periodo = (request.POST.get('periodo') or '').strip()
        if empresa is None:
            return HttpResponseForbidden('No se pudo determinar la empresa del usuario.')

        try:
            resultado = recalcular_consolidado_abierto(empresa=empresa, periodo=periodo, usuario=request.user)
            messages.success(
                request,
                f"Recalculo OK para {resultado.periodo}. Detalles: {resultado.total_detalles_generados}.",
            )
        except (ValueError, ConsolidadoNoExisteError, ConsolidadoBloqueadoError) as exc:
            messages.error(request, str(exc))
        return redirect('core:sindicato_consolidado_historial')


class ConsolidadoSindicatoCerrarPeriodoView(SindicatoTenantMixin, SindicatoRolePermissionMixin, View):
    permiso_ver = 'consolidado_editar'
    permiso_editar = 'consolidado_editar'

    def post(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        periodo = (request.POST.get('periodo') or '').strip()
        if empresa is None:
            return HttpResponseForbidden('No se pudo determinar la empresa del usuario.')

        try:
            consolidado = cerrar_consolidado_periodo(empresa=empresa, periodo=periodo, usuario=request.user)
            messages.success(request, f"Periodo {consolidado.periodo} cerrado correctamente.")
        except (ValueError, ConsolidadoNoExisteError, ConsolidadoBloqueadoError) as exc:
            messages.error(request, str(exc))
        return redirect('core:sindicato_consolidado_historial')


class ConsolidadoSindicatoExportarView(SindicatoTenantMixin, SindicatoRolePermissionMixin, View):
    permiso_ver = 'consolidado_editar'

    def get(self, request, *args, **kwargs):
        empresa = self.get_empresa_usuario()
        consolidado_id = kwargs.get('pk')
        if empresa is None:
            return HttpResponseForbidden('No se pudo determinar la empresa del usuario.')

        try:
            result = exportar_consolidado_excel(
                empresa=empresa,
                consolidado_id=consolidado_id,
                usuario=request.user,
            )
        except ConsolidadoExportacionError as exc:
            messages.error(request, str(exc))
            return redirect('core:sindicato_consolidado_historial')

        response = HttpResponse(
            result.content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{result.filename}"'
        return response


# ==================== AUTENTICACIÓN ====================
class LoginView(View):
    template_name = "core/login.html"
    form_class = EmailAuthenticationForm

    def get_redirect_url(self, user):
        """Retorna la URL de redirección según el rol del usuario"""
        if user.is_superuser or user.groups.filter(name='Admin').exists():
            return "core:dashboard"
        elif user.groups.filter(name='Ejecutivo').exists():
            return "core:ejecutivo_clientes"
        return "core:dashboard"

    def get(self, request):
        if request.user.is_authenticated:
            redirect_url = self.get_redirect_url(request.user)
            return redirect(redirect_url)
        form = self.form_class()
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user, backend="core.backends.EmailBackend")
            redirect_url = self.get_redirect_url(user)
            return redirect(redirect_url)
        return render(request, self.template_name, {"form": form})


# ==================== EJECUTIVO VIEW ====================
class EjecutivoClientesView(LoginRequiredMixin, TemplateView):
    """
    Vista personalizada para ejecutivos.
    Muestra solo los clientes asignados al ejecutivo.
    """
    template_name = "core/ejecutivo_dashboard.html"
    
    def dispatch(self, request, *args, **kwargs):
        # Verifica que sea ejecutivo o vendedor
        if not request.user.groups.filter(name__in=['Ejecutivo', 'Vendedor']).exists():
            if request.user.is_superuser or request.user.groups.filter(name='Admin').exists():
                # Si es admin, redirige al dashboard
                return redirect('core:dashboard')
            # Si no es ni admin ni ejecutivo/vendedor, redirige al login
            return redirect('core:login')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Clientes asignados al ejecutivo
        clientes = Cliente.objects.filter(usuario_asignado=user)
        context['clientes'] = clientes
        context['total_clientes'] = clientes.count()

        # Oportunidades: donde el cliente le pertenece O él es el ejecutivo directo
        oportunidades = Oportunidad.objects.filter(
            Q(usuario=user) | Q(cliente__usuario_asignado=user),
            estado='ABIERTA'
        ).select_related('cliente').distinct()
        context['oportunidades'] = oportunidades
        context['total_oportunidades'] = oportunidades.count()

        # Total líneas de cartera
        context['total_lineas'] = clientes.aggregate(
            total=Sum('q_total_lineas')
        )['total'] or 0

        # Seguimientos pendientes
        context['seguimientos_pendientes'] = Seguimiento.objects.filter(
            Q(oportunidad__usuario=user) | Q(oportunidad__cliente__usuario_asignado=user),
            estado__in=['PENDIENTE', 'EN_PROGRESO']
        ).order_by('fecha_vencimiento').distinct()[:10]
        
        # Gráficos de líneas por cliente
        top_clientes = list(
            clientes.filter(q_total_lineas__gt=0)
            .order_by('-q_total_lineas')
            .values('nombre_empresa', 'q_total_lineas')[:10]
        )
        context['lineas_clientes_labels'] = json.dumps([c['nombre_empresa'] for c in top_clientes])
        context['lineas_clientes_data']   = json.dumps([c['q_total_lineas'] or 0 for c in top_clientes])

        totales = clientes.aggregate(
            total_bam=Sum('bam'),
            total_m2m=Sum('m2m'),
            total_voz=Sum('voz'),
        )
        context['tipos_lineas_labels'] = json.dumps(['BAM', 'M2M', 'VOZ / GPS'])
        context['tipos_lineas_data']   = json.dumps([
            totales['total_bam'] or 0,
            totales['total_m2m'] or 0,
            totales['total_voz'] or 0,
        ])

        context['user_role'] = 'ejecutivo'
        context['nombre_usuario'] = user.get_full_name() or user.email

        return context


# ==================== ADMIN SPANISH VIEW ====================
class AdminRedirectView(View):
    """Redirect to admin with Spanish language set"""
    def get(self, request):
        response = HttpResponseRedirect('/admin/')
        response.set_cookie('django_language', 'es')
        return response


class LogoutView(DjangoLogoutView):
    next_page = "core:login"
    
    def get(self, request, *args, **kwargs):
        """Allow GET requests for logout (in addition to POST)"""
        return self.post(request, *args, **kwargs)
